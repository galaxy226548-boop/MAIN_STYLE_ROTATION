"""Replace legacy ZHAO factor names in generated xlsx outputs.

The script builds a ZHAOxx -> factor_id mapping from
B_factors/reference/factor_done.json, optionally merges a manual override JSON,
then replaces matching tokens in workbook text cells and in xlsx/result paths.

Dry-run is the default. Pass --apply to write workbooks and rename paths.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[1]
DEFAULT_FACTOR_JSON = PROJECT_ROOT / "B_factors" / "reference" / "factor_done.json"

ZHAO_TOKEN_RE = re.compile(r"ZHAO\d+")
EXCEL_SUFFIX = ".xlsx"


@dataclass
class WorkbookChange:
    path: Path
    replacement_count: int = 0
    examples: list[str] = field(default_factory=list)


@dataclass
class RenameChange:
    old_path: Path
    new_path: Path
    path_type: str
    status: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Replace ZHAO factor names in generated xlsx files with factor_id values."
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=PROJECT_ROOT,
        help=f"Project root to scan. Defaults to {PROJECT_ROOT}.",
    )
    parser.add_argument(
        "--factor-json",
        type=Path,
        default=DEFAULT_FACTOR_JSON,
        help=f"factor_done.json path. Defaults to {DEFAULT_FACTOR_JSON}.",
    )
    parser.add_argument(
        "--override-json",
        type=Path,
        help=(
            "Optional manual mapping JSON. Accepts a flat dict like "
            '{"ZHAO05": "X001"} or a list of records with code/factor_id.'
        ),
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually update workbooks and rename paths. Default is dry-run.",
    )
    return parser.parse_args()


def resolve_path(path: Path, root: Path) -> Path:
    if path.is_absolute():
        return path
    return root / path


def load_factor_mapping(factor_json: Path) -> tuple[dict[str, str], dict[str, list[str]]]:
    with factor_json.open("r", encoding="utf-8") as handle:
        data = json.load(handle)

    mapping: dict[str, str] = {}
    missing: dict[str, list[str]] = {}
    conflicts: dict[str, list[str]] = {}

    for sheet_name, sheet in data.get("sheets", {}).items():
        for record in sheet.get("records", []):
            code = str(record.get("code") or "").strip()
            if not ZHAO_TOKEN_RE.fullmatch(code):
                continue

            factor_id = record.get("factor_id")
            if factor_id is None or str(factor_id).strip() == "":
                missing.setdefault(code, []).append(sheet_name)
                continue

            factor_id_text = str(factor_id).strip()
            existing = mapping.get(code)
            if existing and existing != factor_id_text:
                conflicts.setdefault(code, [existing]).append(factor_id_text)
                continue
            mapping[code] = factor_id_text

    if conflicts:
        conflict_text = "; ".join(
            f"{code}: {sorted(set(values))}" for code, values in sorted(conflicts.items())
        )
        raise ValueError(f"Conflicting factor_id mappings in {factor_json}: {conflict_text}")

    return mapping, missing


def normalize_override_records(raw: Any) -> dict[str, str]:
    if isinstance(raw, dict):
        records = raw.items()
    elif isinstance(raw, list):
        records = []
        for item in raw:
            if not isinstance(item, dict):
                raise ValueError("Override list entries must be objects.")
            records.append((item.get("code"), item.get("factor_id")))
    else:
        raise ValueError("Override JSON must be a dict or a list of records.")

    override: dict[str, str] = {}
    for code, factor_id in records:
        code_text = str(code or "").strip()
        factor_id_text = str(factor_id or "").strip()
        if not code_text or not factor_id_text:
            continue
        if not ZHAO_TOKEN_RE.fullmatch(code_text):
            raise ValueError(f"Override code is not a ZHAO token: {code_text}")
        override[code_text] = factor_id_text

    return override


def load_override_mapping(path: Path | None) -> dict[str, str]:
    if path is None:
        return {}
    with path.open("r", encoding="utf-8") as handle:
        return normalize_override_records(json.load(handle))


def replace_text(value: str, mapping: dict[str, str]) -> tuple[str, int]:
    count = 0

    def repl(match: re.Match[str]) -> str:
        nonlocal count
        token = match.group(0)
        replacement = mapping.get(token)
        if replacement is None:
            return token
        count += 1
        return replacement

    return ZHAO_TOKEN_RE.sub(repl, value), count


def scan_workbooks(root: Path) -> list[Path]:
    return sorted(
        path
        for path in root.rglob(f"*ZHAO*{EXCEL_SUFFIX}")
        if path.is_file() and not path.name.startswith("~$")
    )


def replace_workbook_tokens(path: Path, mapping: dict[str, str], apply: bool) -> WorkbookChange:
    workbook = load_workbook(path, data_only=False)
    change = WorkbookChange(path=path)

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or "ZHAO" not in value:
                    continue
                new_value, count = replace_text(value, mapping)
                if count == 0 or new_value == value:
                    continue
                change.replacement_count += count
                if len(change.examples) < 5:
                    change.examples.append(f"{sheet.title}!{cell.coordinate}: {value} -> {new_value}")
                if apply:
                    cell.value = new_value

    if apply and change.replacement_count:
        workbook.save(path)

    return change


def rename_path_text(path: Path, mapping: dict[str, str]) -> Path:
    new_name, count = replace_text(path.name, mapping)
    if count == 0 or new_name == path.name:
        return path
    return path.with_name(new_name)


def collect_rename_changes(workbook_paths: list[Path], mapping: dict[str, str]) -> list[RenameChange]:
    changes: list[RenameChange] = []
    seen: set[Path] = set()

    for path in workbook_paths:
        new_file_path = rename_path_text(path, mapping)
        if new_file_path != path and path not in seen:
            status = "ready"
            if new_file_path.exists():
                status = "collision"
            changes.append(RenameChange(path, new_file_path, "file", status))
            seen.add(path)

        parent = path.parent
        new_parent = rename_path_text(parent, mapping)
        if new_parent != parent and parent not in seen:
            status = "ready"
            if new_parent.exists():
                status = "collision"
            changes.append(RenameChange(parent, new_parent, "dir", status))
            seen.add(parent)

    return changes


def apply_renames(changes: list[RenameChange]) -> None:
    # Rename deeper paths first so file renames happen before their parent dirs.
    ready_changes = [change for change in changes if change.status == "ready"]
    ready_changes.sort(key=lambda change: len(change.old_path.parts), reverse=True)
    for change in ready_changes:
        if not change.old_path.exists():
            change.status = "missing"
            continue
        if change.new_path.exists():
            change.status = "collision"
            continue
        change.old_path.rename(change.new_path)
        change.status = "renamed"


def referenced_zhao_tokens(paths: list[Path]) -> set[str]:
    tokens: set[str] = set()
    for path in paths:
        tokens.update(ZHAO_TOKEN_RE.findall(path.name))
        tokens.update(ZHAO_TOKEN_RE.findall(path.parent.name))
    return tokens


def print_report(
    *,
    apply: bool,
    mapping: dict[str, str],
    json_missing: dict[str, list[str]],
    override: dict[str, str],
    workbook_paths: list[Path],
    workbook_changes: list[WorkbookChange],
    rename_changes: list[RenameChange],
) -> None:
    mode = "APPLY" if apply else "DRY-RUN"
    changed_workbooks = [change for change in workbook_changes if change.replacement_count]
    mapped_tokens_in_paths = sorted(referenced_zhao_tokens(workbook_paths) & set(mapping))
    unresolved_tokens_in_paths = sorted(referenced_zhao_tokens(workbook_paths) - set(mapping))

    print(f"Mode: {mode}")
    print(f"Mapping count: {len(mapping)}")
    print(f"Override count: {len(override)}")
    print(f"Workbook files scanned: {len(workbook_paths)}")
    print(f"Workbooks with cell replacements: {len(changed_workbooks)}")
    print(f"Total cell token replacements: {sum(c.replacement_count for c in workbook_changes)}")
    print(f"Path tokens with mapping: {mapped_tokens_in_paths}")
    print(f"Path tokens without mapping: {unresolved_tokens_in_paths}")

    json_missing_not_overridden = sorted(set(json_missing) - set(override))
    if json_missing_not_overridden:
        print("ZHAO records missing factor_id in factor JSON and not overridden:")
        for code in json_missing_not_overridden:
            print(f"- {code}: sheets={sorted(set(json_missing[code]))}")

    if changed_workbooks:
        print("Workbook replacement details:")
        for change in changed_workbooks:
            print(f"- {change.path}: {change.replacement_count}")
            for example in change.examples:
                print(f"  {example}")

    if rename_changes:
        print("Path rename details:")
        for change in rename_changes:
            print(f"- [{change.status}] {change.path_type}: {change.old_path} -> {change.new_path}")


def main() -> None:
    args = parse_args()
    root = resolve_path(args.root, PROJECT_ROOT).resolve()
    factor_json = resolve_path(args.factor_json, root).resolve()
    override_json = resolve_path(args.override_json, root).resolve() if args.override_json else None

    mapping, json_missing = load_factor_mapping(factor_json)
    override = load_override_mapping(override_json)
    mapping.update(override)

    workbook_paths = scan_workbooks(root)
    workbook_changes = [
        replace_workbook_tokens(path, mapping, apply=args.apply)
        for path in workbook_paths
    ]
    rename_changes = collect_rename_changes(workbook_paths, mapping)

    if args.apply:
        apply_renames(rename_changes)

    print_report(
        apply=args.apply,
        mapping=mapping,
        json_missing=json_missing,
        override=override,
        workbook_paths=workbook_paths,
        workbook_changes=workbook_changes,
        rename_changes=rename_changes,
    )


if __name__ == "__main__":
    main()
