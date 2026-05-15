"""Shared utilities for factor matrix generation and mounting."""

from __future__ import annotations

import json
import sys
import warnings
from pathlib import Path
from typing import Optional
from datetime import datetime, timezone

import numpy as np
import pandas as pd


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[2]
SY_BASELINE_DIR = PROJECT_ROOT / "SY_Baseline"
for import_dir in [PROJECT_ROOT, SY_BASELINE_DIR]:
    if str(import_dir) not in sys.path:
        sys.path.insert(0, str(import_dir))

from Config import Config


data_df_path = PROJECT_ROOT / "A_data" / "prepared_data" / "data_df.parquet"
market_df_path = PROJECT_ROOT / "A_data" / "prepared_data" / "market_df.parquet"
benchmark_positions_parquet_path = PROJECT_ROOT / "C_positions" / "reference" / "benchmark_positions.parquet"
benchmark_positions_csv_path = PROJECT_ROOT / "C_positions" / "reference" / "benchmark_positions.csv"
factor_done_path = PROJECT_ROOT / "B_factors" / "reference" / "factor_done.json"
record_all_path = PROJECT_ROOT / "B_factors" / "reference" / "record_all.json"
output_dir = PROJECT_ROOT / "B_factors" / "output"
prepared_data_dir = PROJECT_ROOT / "A_data" / "prepared_data"
data_inventory_path = PROJECT_ROOT / "A_data" / "reference" / "data_inventory_A.json"

Config.DATA_DIR = prepared_data_dir

if not data_df_path.exists():
    raise FileNotFoundError(f"找不到 data_df 文件：{data_df_path}")

if not market_df_path.exists():
    raise FileNotFoundError(f"找不到 market_df 文件：{market_df_path}")


PREPARED_SERIES_MAP = {
    ("DebtData.xlsx", "中国:金融机构:新增人民币贷款:中长期:当月值"): (
        "macro_monthly.parquet",
        "中国:金融机构:新增人民币贷款:中长期:当月值",
    ),
    ("D_国债到期收益率_CN_020104_260409.xlsx", "中债国债到期收益率:1年"): (
        "rate_daily.parquet",
        "中债国债到期收益率:1年",
    ),
    ("D_国债收益率_US_530430_260324.xlsx", "美国:国债收益率:2年"): (
        "rate_daily.parquet",
        "美国:国债收益率:2年",
    ),
    ("规模以上工业 招证资配.xlsx", "中国:利润总额:规模以上工业企业:累计值"): (
        "macro_monthly.parquet",
        "中国:利润总额:规模以上工业企业:累计值",
    ),
    ("规模以上工业 招证资配.xlsx", "中国:产成品存货:规模以上工业企业:同比"): (
        "macro_monthly.parquet",
        "中国:产成品存货:规模以上工业企业:同比",
    ),
    ("公共预算支出.xlsx", "中国:一般公共预算支出:当月同比(1-2月合并)"): (
        "macro_monthly.parquet",
        "中国:一般公共预算支出:当月同比(1-2月合并)",
    ),
    ("日频汇率.xlsx", "中间价:美元兑人民币"): (
        "exchange_rate_daily.parquet",
        "中间价:美元兑人民币",
    ),
}

_PREPARED_TABLE_CACHE: dict[str, pd.DataFrame] = {}
_DATA_INVENTORY_CACHE: dict[str, object] | None = None


def load_default_data() -> tuple[pd.DataFrame, pd.DataFrame]:
    data_df = pd.read_parquet(data_df_path)
    market_df = pd.read_parquet(market_df_path)
    data_df.index = pd.to_datetime(data_df.index)
    market_df.index = pd.to_datetime(market_df.index)
    return data_df.sort_index(), market_df.sort_index()


def load_benchmark_index() -> pd.DatetimeIndex:
    if benchmark_positions_parquet_path.exists():
        benchmark_df = pd.read_parquet(benchmark_positions_parquet_path)
        benchmark_index = pd.to_datetime(benchmark_df.index)
    elif benchmark_positions_csv_path.exists():
        benchmark_df = pd.read_csv(benchmark_positions_csv_path)
        if "date" not in benchmark_df.columns:
            raise KeyError(f"{benchmark_positions_csv_path} must contain a date column")
        benchmark_index = pd.to_datetime(benchmark_df["date"])
    else:
        raise FileNotFoundError(
            "找不到 benchmark positions 文件："
            f"{benchmark_positions_parquet_path} 或 {benchmark_positions_csv_path}"
        )

    benchmark_index = pd.DatetimeIndex(benchmark_index).sort_values()
    if benchmark_index.has_duplicates:
        duplicated = benchmark_index[benchmark_index.duplicated()].unique()
        raise ValueError(f"benchmark positions 日期存在重复：{duplicated[:5].tolist()}")
    return benchmark_index


def load_data_inventory() -> dict[str, object]:
    global _DATA_INVENTORY_CACHE
    if _DATA_INVENTORY_CACHE is None:
        if not data_inventory_path.exists():
            raise FileNotFoundError(f"找不到 data_inventory_A.json：{data_inventory_path}")
        _DATA_INVENTORY_CACHE = json.loads(data_inventory_path.read_text(encoding="utf-8"))
    return _DATA_INVENTORY_CACHE


def _normalize_clean_data_name(clean_data: str) -> str:
    clean_data = str(clean_data)
    if clean_data in {"none", "", "nan"}:
        return clean_data
    if Path(clean_data).suffix:
        return clean_data
    return clean_data + ".parquet"


def validate_prepared_mapping() -> None:
    inventory = load_data_inventory()
    clean_data_values = set()
    for sheet_meta in inventory.get("sheets", {}).values():
        for record in sheet_meta.get("records", []):
            clean_data = record.get("clean_data")
            if clean_data:
                clean_data_values.add(_normalize_clean_data_name(clean_data))

    for table_name, _column_name in PREPARED_SERIES_MAP.values():
        table_name = _normalize_clean_data_name(table_name)
        table_path = prepared_data_dir / table_name
        if not table_path.exists():
            raise FileNotFoundError(f"映射到的 prepared 文件不存在：{table_path}")
        if table_name not in clean_data_values:
            print(f"Warning: {table_name} 未在 data_inventory_A.json 的 clean_data 中出现，仍按 prepared 文件读取。")


def _prepared_table_path(table_name: str) -> Path:
    table_name = _normalize_clean_data_name(table_name)
    path = prepared_data_dir / table_name
    if not path.exists():
        raise FileNotFoundError(f"找不到 prepared 数据文件：{path}")
    if path.suffix == ".feather":
        print(f"需要读取 feather 文件：{path}")
        raise RuntimeError("feather 不是 B 模块默认数据源；请确认该因子是否非用 feather 不可。")
    return path


def load_prepared_table(table_name: str) -> pd.DataFrame:
    table_name = _normalize_clean_data_name(table_name)
    if table_name not in _PREPARED_TABLE_CACHE:
        path = _prepared_table_path(table_name)
        if path.suffix == ".parquet":
            df = pd.read_parquet(path)
        elif path.suffix in {".xlsx", ".xls"}:
            df = pd.read_excel(path)
        else:
            raise ValueError(f"不支持的 prepared 数据格式：{path}")
        _PREPARED_TABLE_CACHE[table_name] = df
    return _PREPARED_TABLE_CACHE[table_name].copy()


def read_prepared_series(table_name: str, column_name: str) -> pd.Series:
    df = load_prepared_table(table_name)
    if column_name not in df.columns:
        raise KeyError(
            f"{column_name} not found in prepared table {table_name}. "
            f"Available columns: {list(df.columns)}"
        )

    if isinstance(df.index, pd.DatetimeIndex):
        index = pd.to_datetime(df.index)
    else:
        date_col = next((col for col in ["date", "日期", "TRADE_DT", "Trddt"] if col in df.columns), None)
        if date_col is None:
            raise KeyError(f"prepared table {table_name} 没有 datetime index，也找不到日期列")
        index = pd.to_datetime(df[date_col], errors="coerce")

    s = pd.Series(pd.to_numeric(df[column_name], errors="coerce").to_numpy(), index=index, name=column_name)
    s = s[s.index.notna()].sort_index()
    return s[~s.index.duplicated(keep="last")]


def clean_macro_table(df: pd.DataFrame, nation: str = "美国", indi: str = "PMI") -> pd.DataFrame:
    out = df.copy()
    out = out[out["国家/地区"] == nation]
    out = out[out["指标名称"].astype(str).str.contains(indi, na=False)]

    keep_cols = [
        out.columns[2],
        out.columns[3],
        out.columns[4],
        out.columns[5],
        out.columns[7],
        out.columns[8],
        out.columns[9],
        out.columns[-1],
    ]
    out = out[keep_cols].copy()
    out.columns = ["date", "time", "nation", "indicator", "prev", "forecast", "actual", "file_ym"]
    out = out.set_index("date").sort_index(ascending=True)
    out = out.drop_duplicates(
        subset=["time", "nation", "indicator", "prev", "forecast", "actual"],
        keep="last",
    )
    return out


def _find_data_file(file_name):
    file_name = str(file_name)
    if file_name.lower().endswith(".feather"):
        candidate = prepared_data_dir / file_name
        print(f"需要读取 feather 文件：{candidate}")
        raise RuntimeError("feather 不是 B 模块默认数据源；请确认该因子是否非用 feather 不可。")

    direct = prepared_data_dir / file_name
    if direct.exists():
        return direct
    candidates = sorted(prepared_data_dir.rglob(file_name))
    if not candidates and not file_name.lower().endswith((".xlsx", ".xls")):
        candidates = sorted(prepared_data_dir.rglob(file_name + ".xlsx"))
    if not candidates:
        fallback_root = PROJECT_ROOT / "A_data" / "data"
        candidates = sorted(fallback_root.rglob(file_name))
        if not candidates and not file_name.lower().endswith((".xlsx", ".xls")):
            candidates = sorted(fallback_root.rglob(file_name + ".xlsx"))
    if not candidates:
        raise FileNotFoundError(f"Cannot find data file under {prepared_data_dir}: {file_name}")
    if len(candidates) > 1:
        print(f"Multiple files matched {file_name}, using: {candidates[0]}")
    return candidates[0]


def _as_numeric(series, percent_hint=False):
    text = series.astype(str).str.strip()
    has_percent_sign = text.str.contains("%", regex=False, na=False).any()
    numeric = pd.to_numeric(
        text.str.replace("%", "", regex=False).str.replace(",", "", regex=False),
        errors="coerce",
    )
    if percent_hint or has_percent_sign:
        numeric = numeric / 100
    return numeric


def _read_indicator_series(file_name, value_col, sheet_name=0):
    prepared_mapping = PREPARED_SERIES_MAP.get((str(file_name), str(value_col)))
    if prepared_mapping is not None:
        table_name, prepared_col = prepared_mapping
        return read_prepared_series(table_name, prepared_col)

    path = _find_data_file(file_name)
    if path.suffix == ".feather":
        print(f"需要读取 feather 文件：{path}")
        raise RuntimeError("feather 不是 B 模块默认数据源；请确认该因子是否非用 feather 不可。")
    raw = pd.read_excel(path, sheet_name=sheet_name)
    date_col = raw.columns[0]
    unit_mask = raw[date_col].astype(str).str.strip().eq("单位")
    percent_hint = False
    if value_col in raw.columns and unit_mask.any():
        percent_hint = raw.loc[unit_mask, value_col].astype(str).str.contains("%", regex=False, na=False).any()
    out = raw.copy()
    out[date_col] = pd.to_datetime(out[date_col], errors="coerce")
    out = out[out[date_col].notna()].copy()
    out = out.set_index(date_col).sort_index(ascending=True)
    if value_col not in out.columns:
        raise KeyError(f"{value_col} not found in {path}. Available columns: {list(out.columns)}")
    s = _as_numeric(out[value_col], percent_hint=percent_hint)
    s.name = value_col
    return s[~s.index.duplicated(keep="last")].sort_index()


def _month_aggregate(series, how="average"):
    s = series.dropna().copy().sort_index()
    grouped = s.groupby(s.index.to_period("M"))
    if how in ["average", "mean"]:
        out = grouped.mean()
    elif how == "last":
        out = grouped.last()
    elif how == "sum":
        out = grouped.sum()
    else:
        raise ValueError(f"Unsupported aggregate method: {how}")
    last_dates = grouped.apply(lambda x: x.index[-1])
    out.index = last_dates.values
    return out


def _rolling_quantile_rank_year(series, year=5):
    s = series.dropna().sort_index()
    dates = s.index
    out = []
    left = 0
    if len(s) == 0:
        return pd.Series(dtype="float64")
    first_date = dates[0]
    for right in range(len(s)):
        end_date = dates[right]
        start_date = end_date - pd.DateOffset(years=year)
        if first_date > start_date:
            out.append(np.nan)
            continue
        while left < right and dates[left] < start_date:
            left += 1
        window = s.iloc[left:right + 1]
        out.append(window.rank(pct=True).iloc[-1])
    return pd.Series(out, index=s.index)


def data_diff(series: pd.Series) -> pd.Series:
    s = series.copy().sort_index()
    return s - s.shift(1)


def data_yoy(series: pd.Series) -> pd.Series:
    s = series.copy().sort_index()
    return s / s.shift(1) - 1


def calc_rolling_zscore(series: pd.Series, window: int, min_periods: Optional[int] = None) -> pd.Series:
    if min_periods is None:
        min_periods = window // 2
    s = series.astype("float64").sort_index()
    rolling_mean = s.rolling(window=window, min_periods=min_periods).mean()
    rolling_std = s.rolling(window=window, min_periods=min_periods).std()
    return (s - rolling_mean) / rolling_std


def calc_llt(series: pd.Series, d: int = 30) -> pd.Series:
    s = series.astype("float64").copy().sort_index()
    llt_series = pd.Series(index=s.index, dtype="float64")
    alpha = 2 / (d + 1)
    started = False

    for i in range(2, len(s)):
        x_t = s.iloc[i]
        x_t1 = s.iloc[i - 1]
        x_t2 = s.iloc[i - 2]
        if pd.isna(x_t) or pd.isna(x_t1) or pd.isna(x_t2):
            llt_series.iloc[i] = np.nan
            continue

        if not started:
            llt_series.iloc[i - 1] = x_t1
            llt_series.iloc[i] = x_t
            started = True
            continue

        llt_t1 = llt_series.iloc[i - 1]
        llt_t2 = llt_series.iloc[i - 2]
        if pd.isna(llt_t1) or pd.isna(llt_t2):
            llt_series.iloc[i - 1] = x_t1
            llt_series.iloc[i] = x_t
            continue

        llt_series.iloc[i] = (
            (alpha - alpha ** 2 / 4) * x_t
            + (alpha ** 2 / 2) * x_t1
            - (alpha - 3 * alpha ** 2 / 4) * x_t2
            + 2 * (1 - alpha) * llt_t1
            - (1 - alpha) ** 2 * llt_t2
        )

    return llt_series


def _load_macro_all():
    return load_prepared_table("macro.parquet")


def _macro_monthly_fallback_series(keyword: str, value_col: str = "今值") -> pd.Series:
    if value_col not in {"今值", "actual"}:
        raise ValueError(f"macro_monthly fallback only supports current values, got value_col={value_col!r}")

    macro_monthly = load_prepared_table("macro_monthly.parquet")
    keyword_text = str(keyword)
    candidates = [col for col in macro_monthly.columns if keyword_text in str(col)]

    if not candidates:
        simplified = (
            keyword_text
            .replace("月", "")
            .replace("(%)", "")
            .replace("(亿美元)", "")
            .replace(":当月值", "")
        )
        candidates = [col for col in macro_monthly.columns if simplified and simplified in str(col)]

    if not candidates:
        raise ValueError(f"No China macro rows matched keyword={keyword!r}, and no macro_monthly fallback column matched")
    if len(candidates) > 1:
        exact = [col for col in candidates if str(col).endswith(keyword_text) or str(col) == keyword_text]
        if len(exact) == 1:
            candidates = exact
        else:
            raise ValueError(f"macro_monthly fallback for keyword={keyword!r} matched multiple columns: {candidates}")

    return read_prepared_series("macro_monthly.parquet", candidates[0])


def _fill_missing_current_from_next_prev(
    out: pd.DataFrame,
    date_col: str,
    value_col: str,
    keyword: str,
    current_values: pd.Series,
    prev_values: pd.Series | None,
    duplicate_message_prefix: str,
) -> pd.Series:
    work = pd.DataFrame(
        {
            "date": pd.to_datetime(out[date_col], errors="coerce").to_numpy(),
            "actual": pd.Series(current_values, index=out.index).to_numpy(dtype="float64"),
        }
    )
    if prev_values is not None:
        work["prev"] = pd.Series(prev_values, index=out.index).to_numpy(dtype="float64")
    else:
        work["prev"] = np.nan

    work = work[work["date"].notna()].copy()
    dup_count = int(work["date"].duplicated(keep=False).sum())
    if dup_count > 0:
        print(f"{duplicate_message_prefix} keyword {keyword!r} matched {dup_count} duplicate-date rows; keeping the last row per date.")
        work = work[~work["date"].duplicated(keep="last")].copy()
    work = work.sort_values("date")

    if value_col not in {"今值", "actual"}:
        return pd.Series(work["actual"].to_numpy(), index=work["date"], name=keyword).sort_index()

    actual = pd.Series(work["actual"].to_numpy(), index=work["date"], name=keyword)
    next_prev = pd.Series(work["prev"].to_numpy(), index=work["date"]).shift(-1)
    next_date = pd.Series(work["date"].shift(-1).to_numpy(), index=work["date"])

    comparable = actual.notna() & next_prev.notna()
    mismatch = comparable & ~np.isclose(actual, next_prev, rtol=1e-9, atol=1e-12)
    if mismatch.any():
        rows = pd.DataFrame(
            {
                "date": actual.index[mismatch],
                "actual": actual.loc[mismatch].to_numpy(),
                "next_date": next_date.loc[mismatch].to_numpy(),
                "next_prev": next_prev.loc[mismatch].to_numpy(),
            }
        )
        warnings.warn(
            f"Macro keyword {keyword!r} has {len(rows)} rows where current value does not match "
            "the next row's previous value:\n"
            f"{rows.head(10).to_string(index=False)}",
            stacklevel=2,
        )

    fill_mask = actual.isna() & next_prev.notna()
    if fill_mask.any():
        actual.loc[fill_mask] = next_prev.loc[fill_mask]

    remaining_missing = actual[actual.isna()]
    if not remaining_missing.empty:
        dates = [dt.strftime("%Y-%m-%d") for dt in remaining_missing.index[:10]]
        suffix = "" if len(remaining_missing) <= 10 else f" ... total={len(remaining_missing)}"
        warnings.warn(
            f"Macro keyword {keyword!r} still has missing current values after next-prev fill: "
            f"{dates}{suffix}",
            stacklevel=2,
        )

    return actual.sort_index()


def _load_china_macro_series(keyword, value_col="今值", required_contains=None, exclude_contains=None):
    macro = _load_macro_all()
    date_col = "日期" if "日期" in macro.columns else macro.columns[2]
    nation_col = "国家/地区" if "国家/地区" in macro.columns else macro.columns[4]
    indicator_col = "指标名称" if "指标名称" in macro.columns else macro.columns[5]
    mask = macro[nation_col].eq("中国")
    mask &= macro[indicator_col].astype(str).str.contains(keyword, na=False, regex=False)
    if required_contains is not None:
        required_list = [required_contains] if isinstance(required_contains, str) else list(required_contains)
        for item in required_list:
            mask &= macro[indicator_col].astype(str).str.contains(item, na=False, regex=False)
    if exclude_contains is not None:
        exclude_list = [exclude_contains] if isinstance(exclude_contains, str) else list(exclude_contains)
        for item in exclude_list:
            mask &= ~macro[indicator_col].astype(str).str.contains(item, na=False, regex=False)
    out = macro.loc[mask].copy()
    if out.empty:
        return _macro_monthly_fallback_series(keyword, value_col=value_col)
    percent_hint = (
        out[indicator_col].astype(str).str.contains("%", regex=False, na=False).any()
        or out[value_col].astype(str).str.contains("%", regex=False, na=False).any()
    )
    out[date_col] = pd.to_datetime(out[date_col], errors="coerce")
    out = out[out[date_col].notna()].copy()
    sort_cols = [x for x in [date_col, "来源文件", "来源sheet", "文件年月"] if x in out.columns]
    out = out.sort_values(sort_cols, na_position="first")
    current_values = _as_numeric(out[value_col], percent_hint=percent_hint)
    prev_values = _as_numeric(out["前值"], percent_hint=percent_hint) if "前值" in out.columns else None
    s = _fill_missing_current_from_next_prev(
        out,
        date_col=date_col,
        value_col=value_col,
        keyword=keyword,
        current_values=current_values,
        prev_values=prev_values,
        duplicate_message_prefix="Macro",
    )
    return s.sort_index()


def _load_china_macro_level_series(keyword: str, value_col: str = "今值") -> pd.Series:
    """Load China macro calendar values as raw levels, without percent-style /100 scaling.

    Use this for level/diffusion-index indicators such as PMI. Some macro indicator names
    include "(%)" even though the usable value should remain 50.2 rather than 0.502.
    """
    macro = _load_macro_all()
    date_col = "日期" if "日期" in macro.columns else macro.columns[2]
    nation_col = "国家/地区" if "国家/地区" in macro.columns else macro.columns[4]
    indicator_col = "指标名称" if "指标名称" in macro.columns else macro.columns[5]
    mask = macro[nation_col].eq("中国")
    mask &= macro[indicator_col].astype(str).str.contains(keyword, na=False, regex=False)
    out = macro.loc[mask].copy()
    if out.empty:
        raise ValueError(f"macro.parquet 中找不到中国宏观指标：{keyword!r}")
    out[date_col] = pd.to_datetime(out[date_col], errors="coerce")
    out = out[out[date_col].notna()].copy()
    sort_cols = [x for x in [date_col, "来源文件", "来源sheet", "文件年月"] if x in out.columns]
    out = out.sort_values(sort_cols, na_position="first")
    current_values = pd.to_numeric(out[value_col], errors="coerce")
    prev_values = pd.to_numeric(out["前值"], errors="coerce") if "前值" in out.columns else None
    s = _fill_missing_current_from_next_prev(
        out,
        date_col=date_col,
        value_col=value_col,
        keyword=keyword,
        current_values=current_values,
        prev_values=prev_values,
        duplicate_message_prefix="Macro level",
    )
    return s.sort_index()


def _rolling_sum_ratio_minus_one(series, window=12, shift=11):
    rolling_sum = series.dropna().sort_index().rolling(window=window, min_periods=window).sum().dropna()
    division = rolling_sum / rolling_sum.shift(shift) - 1
    return division.dropna()


def _YoY(monthly_series):
    s = monthly_series.dropna().sort_index()
    return s / s.shift(1) - 1


def _register_factor(
    _raw_factor_df: pd.DataFrame,
    normalized_factor_df: pd.DataFrame,
    raw_col: str,
    raw_factor_series: pd.Series,
    normalized_factor_series: Optional[pd.Series] = None,
) -> None:
    if not raw_col.endswith("_raw"):
        raise ValueError(f"{raw_col} must end with _raw")
    factor_col = raw_col[:-4]

    if normalized_factor_series is None:
        normalized_factor_series = raw_factor_series

    normalized_factor = normalized_factor_series.copy().sort_index()
    normalized_factor.index = pd.to_datetime(normalized_factor.index)
    normalized_factor = normalized_factor[~normalized_factor.index.duplicated(keep="last")]
    missing_index = normalized_factor.index.difference(normalized_factor_df.index)
    for dt in missing_index:
        normalized_factor_df.loc[dt, :] = np.nan
    if len(missing_index) > 0:
        normalized_factor_df.sort_index(inplace=True)
    normalized_factor_df[factor_col] = normalized_factor.reindex(normalized_factor_df.index).astype("float64")

    print(
        f"{raw_col} registered:",
        "source_factor_non_na=", int(normalized_factor_df[factor_col].notna().sum()),
        "first=", normalized_factor_df[factor_col].first_valid_index(),
        "last=", normalized_factor_df[factor_col].last_valid_index(),
    )


def pmi_yoy_chain(
    df_PMI: pd.DataFrame,
    value_col: str = "actual",
    scale: float = 0.01,
    base_level: float = 100.0,
    drop_dup_keep: str = "last",
) -> pd.DataFrame:
    out = df_PMI.copy()
    out = out.sort_index()
    out = out[~out.index.duplicated(keep=drop_dup_keep)].copy()
    out["PMI"] = pd.to_numeric(out[value_col], errors="coerce")
    out["mom_proxy"] = (out["PMI"] - 50.0) * scale
    growth_factor = 1.0 + out["mom_proxy"]
    out["pseudo_level"] = base_level * growth_factor.cumprod()
    out["yoy_chain"] = out["pseudo_level"] / out["pseudo_level"].shift(12) - 1.0
    return out


def normalize_trade_dt(series):
    text = series.astype("string").str.strip().str.replace(r"\.0$", "", regex=True)
    is_yyyymmdd = text.str.fullmatch(r"\d{8}").fillna(False)
    dt = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    dt.loc[is_yyyymmdd] = pd.to_datetime(text.loc[is_yyyymmdd], format="%Y%m%d", errors="coerce")
    dt.loc[~is_yyyymmdd] = pd.to_datetime(text.loc[~is_yyyymmdd], errors="coerce")
    return dt.dt.normalize()


def calc_qrd(x: pd.Series) -> float:
    x = x.dropna()
    if len(x) < 10:
        return np.nan
    q10 = x.quantile(0.10)
    q25 = x.quantile(0.25)
    q75 = x.quantile(0.75)
    q90 = x.quantile(0.90)
    denominator = q90 - q10
    if denominator == 0:
        return np.nan
    return (q75 - q25) / denominator


def _normalize_signal_type(signal_type: object) -> str:
    value = str(signal_type).strip().lower()
    if value == "stsate":
        value = "state"
    if value not in {"state", "event"}:
        raise ValueError(f"Unsupported signal_type: {signal_type!r}")
    return value


def load_factor_metadata(factor_cols: list[str]) -> tuple[dict[str, dict[str, object]], list[dict[str, object]]]:
    if not factor_done_path.exists():
        raise FileNotFoundError(f"找不到 factor_done.json：{factor_done_path}")

    payload = json.loads(factor_done_path.read_text(encoding="utf-8"))
    records = []
    for sheet_name, sheet_meta in payload.get("sheets", {}).items():
        for record in sheet_meta.get("records", []):
            if record.get("code") is not None:
                item = dict(record)
                item["_sheet"] = sheet_name
                records.append(item)

    by_code: dict[str, list[dict[str, object]]] = {}
    for record in records:
        by_code.setdefault(str(record["code"]), []).append(record)

    metadata: dict[str, dict[str, object]] = {}
    missing_bar_defaults: list[dict[str, object]] = []
    for factor_col in factor_cols:
        matches = by_code.get(factor_col, [])
        if not matches:
            raise KeyError(f"factor_done.json 中找不到因子 code={factor_col}")
        if len(matches) > 1:
            done_matches = [x for x in matches if x.get("progress") in {"done", "completed"}]
            matches = done_matches or matches
        if len(matches) > 1:
            raise ValueError(f"factor_done.json 中 code={factor_col} 有多条候选，无法唯一匹配")

        record = matches[0]
        signal_type = _normalize_signal_type(record.get("signal_type"))
        bar = record.get("bar")
        used_default_bar = bar is None
        if used_default_bar:
            bar = 0
            missing_bar_defaults.append(
                {
                    "code": factor_col,
                    "factor": record.get("factor"),
                    "signal_type": signal_type,
                    "progress": record.get("progress"),
                    "default_bar": bar,
                }
            )
        metadata[factor_col] = {
            "signal_type": signal_type,
            "bar": float(bar),
            "factor": record.get("factor"),
            "progress": record.get("progress"),
        }

    return metadata, missing_bar_defaults


def _metadata_from_records(
    records: list[dict[str, object]],
    factor_cols: list[str],
    key_name: str,
    source_label: str,
) -> tuple[dict[str, dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    by_key: dict[str, list[dict[str, object]]] = {}
    for record in records:
        key_value = record.get(key_name)
        if key_value is not None:
            by_key.setdefault(str(key_value), []).append(record)

    metadata: dict[str, dict[str, object]] = {}
    missing_bar_defaults: list[dict[str, object]] = []
    selected_records: list[dict[str, object]] = []

    for factor_col in factor_cols:
        matches = by_key.get(factor_col, [])
        if not matches:
            raise KeyError(f"{source_label} 中找不到 {key_name}={factor_col}")
        if len(matches) > 1:
            done_matches = [x for x in matches if x.get("progress") in {"done", "completed"}]
            matches = done_matches or matches
        if len(matches) > 1:
            raise ValueError(f"{source_label} 中 {key_name}={factor_col} 有多条候选，无法唯一匹配")

        record = matches[0]
        signal_type = _normalize_signal_type(record.get("signal_type"))
        bar = record.get("bar")
        used_default_bar = bar is None
        if used_default_bar:
            bar = 0
            missing_bar_defaults.append(
                {
                    key_name: factor_col,
                    "factor": record.get("factor"),
                    "signal_type": signal_type,
                    "progress": record.get("progress"),
                    "default_bar": bar,
                }
            )
        metadata[factor_col] = {
            "signal_type": signal_type,
            "bar": float(bar),
            "factor": record.get("factor"),
            "progress": record.get("progress"),
        }
        selected_records.append(dict(record))

    return metadata, missing_bar_defaults, selected_records


def load_factor_done_factor_id_metadata(
    factor_cols: list[str],
) -> tuple[dict[str, dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    if not factor_done_path.exists():
        raise FileNotFoundError(f"找不到 factor_done.json：{factor_done_path}")

    payload = json.loads(factor_done_path.read_text(encoding="utf-8"))
    records = []
    source_file = str(factor_done_path.relative_to(PROJECT_ROOT))
    for sheet_name, sheet_meta in payload.get("sheets", {}).items():
        for record in sheet_meta.get("records", []):
            if record.get("factor_id") is not None:
                item = dict(record)
                item["_source_file"] = source_file
                item["_source_sheet"] = sheet_name
                records.append(item)

    return _metadata_from_records(records, factor_cols, "factor_id", "factor_done.json")


def save_generated_factor_records(
    records: list[dict[str, object]],
    output_prefix: str,
    generated_path: Path | str | None = None,
) -> Path:
    path = Path(generated_path) if generated_path is not None else output_dir / "factor_generated.json"
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        existing_payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(existing_payload, list):
            existing_records = existing_payload
        else:
            existing_records = existing_payload.get("records", [])
    else:
        existing_records = []

    generated_at = datetime.now(timezone.utc).isoformat()

    def record_key(record: dict[str, object]) -> tuple[str, str, str, str, str]:
        source_file = str(record.get("_source_file") or "")
        source_sheet = str(record.get("_source_sheet") or record.get("_sheet") or "")
        output_prefix_key = str(record.get("_generated_output_prefix") or "")
        factor_id = record.get("factor_id")
        if factor_id is not None:
            return source_file, source_sheet, output_prefix_key, "factor_id", str(factor_id)
        return source_file, source_sheet, output_prefix_key, "code", str(record.get("code") or "")

    merged: dict[tuple[str, str, str, str, str], dict[str, object]] = {}
    for record in existing_records:
        if isinstance(record, dict):
            merged[record_key(record)] = dict(record)

    for record in records:
        item = dict(record)
        item["_generated_output_prefix"] = output_prefix
        item["_generated_at"] = generated_at
        merged[record_key(item)] = item

    payload = {
        "generated_at": generated_at,
        "records": list(merged.values()),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def load_record_all_factor_metadata(
    paper_id: str,
    factor_cols: list[str],
) -> tuple[dict[str, dict[str, object]], list[dict[str, object]]]:
    metadata, missing_bar_defaults, _selected_records = load_record_all_factor_metadata_with_records(
        paper_id,
        factor_cols,
    )
    return metadata, missing_bar_defaults


def load_record_all_factor_metadata_with_records(
    paper_id: str,
    factor_cols: list[str],
) -> tuple[dict[str, dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    if not record_all_path.exists():
        raise FileNotFoundError(f"找不到 record_all.json：{record_all_path}")

    payload = json.loads(record_all_path.read_text(encoding="utf-8"))
    records = []
    for sheet_name, sheet_meta in payload.get("sheets", {}).items():
        for record in sheet_meta.get("records", []):
            if record.get("paper_id") == paper_id and record.get("factor_id") is not None:
                item = dict(record)
                item["_sheet"] = sheet_name
                records.append(item)

    by_factor_id: dict[str, list[dict[str, object]]] = {}
    for record in records:
        by_factor_id.setdefault(str(record["factor_id"]), []).append(record)

    metadata: dict[str, dict[str, object]] = {}
    missing_bar_defaults: list[dict[str, object]] = []
    selected_records: list[dict[str, object]] = []

    for factor_col in factor_cols:
        matches = by_factor_id.get(factor_col, [])
        if not matches:
            raise KeyError(f"record_all.json 中 paper_id={paper_id} 找不到 factor_id={factor_col}")
        if len(matches) > 1:
            raise ValueError(f"record_all.json 中 factor_id={factor_col} 有多条候选，无法唯一匹配")

        record = matches[0]
        signal_type = _normalize_signal_type(record.get("signal_type"))
        bar = record.get("bar")
        used_default_bar = bar is None
        if used_default_bar:
            bar = 0
            missing_bar_defaults.append(
                {
                    "code": factor_col,
                    "factor": record.get("factor"),
                    "signal_type": signal_type,
                    "progress": record.get("progress"),
                    "default_bar": bar,
                }
            )
        metadata[factor_col] = {
            "signal_type": signal_type,
            "bar": float(bar),
            "factor": record.get("factor"),
            "progress": record.get("progress"),
        }
        selected_records.append(dict(record))

    return metadata, missing_bar_defaults, selected_records


def mount_factor_source_frame(
    factor_source_df: pd.DataFrame,
    market_df: pd.DataFrame,
    benchmark_index: pd.DatetimeIndex,
    metadata: dict[str, dict[str, object]],
    track_col: str = "track_id",
) -> pd.DataFrame:
    factor_source_df = factor_source_df.copy().sort_index()
    market_df = market_df.copy().sort_index()
    factor_source_df.index = pd.to_datetime(factor_source_df.index)
    market_df.index = pd.to_datetime(market_df.index)
    benchmark_index = pd.DatetimeIndex(benchmark_index).sort_values()

    if track_col not in market_df.columns:
        raise KeyError(f"event factors require {track_col} in market_df")

    mounted_df = pd.DataFrame(index=market_df.index, columns=factor_source_df.columns, dtype="float64")

    state_cols = [col for col in factor_source_df.columns if metadata[col]["signal_type"] == "state"]
    if state_cols:
        state_values = factor_source_df[state_cols].ffill()
        lookup_index = pd.DatetimeIndex(market_df.index) - pd.Timedelta(nanoseconds=1)
        mounted_state = state_values.reindex(lookup_index, method="ffill")
        mounted_state.index = market_df.index
        mounted_df.loc[:, state_cols] = mounted_state

    event_cols = [col for col in factor_source_df.columns if metadata[col]["signal_type"] == "event"]
    if event_cols:
        track_values = sorted(pd.Series(market_df[track_col]).dropna().unique())
        track_dates = {
            track_id: pd.DatetimeIndex(market_df.index[market_df[track_col] == track_id]).sort_values()
            for track_id in track_values
        }
        for factor_col in event_cols:
            raw_events = factor_source_df.loc[factor_source_df[factor_col].notna(), factor_col].sort_index()
            for event_date, raw_value in raw_events.items():
                event_date = pd.Timestamp(event_date)
                for candidate_dates in track_dates.values():
                    future_dates = candidate_dates[candidate_dates > event_date]
                    if len(future_dates) > 0:
                        mounted_df.loc[future_dates[0], factor_col] = raw_value

    return mounted_df.reindex(benchmark_index).astype("float64")


def build_signal_ls_df(
    mounted_factor_df: pd.DataFrame,
    metadata: dict[str, dict[str, object]],
) -> pd.DataFrame:
    signal_ls_df = pd.DataFrame(index=mounted_factor_df.index, columns=mounted_factor_df.columns, dtype="float64")

    for factor_col in mounted_factor_df.columns:
        factor = mounted_factor_df[factor_col]
        bar = float(metadata[factor_col]["bar"])
        signal_type = str(metadata[factor_col]["signal_type"])

        if signal_type == "state":
            signal = pd.Series(0.0, index=factor.index)
            signal.loc[factor > bar] = 1.0
            signal.loc[factor < -bar] = -1.0
        elif signal_type == "event":
            signal = pd.Series(np.nan, index=factor.index, dtype="float64")
            has_event = factor.notna()
            signal.loc[has_event & (factor > bar)] = 1.0
            signal.loc[has_event & (factor < -bar)] = -1.0
            signal.loc[has_event & factor.between(-bar, bar)] = 0.0
        else:
            raise ValueError(f"Unsupported signal_type for {factor_col}: {signal_type!r}")

        signal_ls_df[factor_col] = signal

    return signal_ls_df


def build_threshold_signal_ls_df(
    mounted_factor_df: pd.DataFrame,
    metadata: dict[str, dict[str, object]],
) -> pd.DataFrame:
    signal_ls_df = pd.DataFrame(index=mounted_factor_df.index, columns=mounted_factor_df.columns, dtype="float64")

    for factor_col in mounted_factor_df.columns:
        factor = mounted_factor_df[factor_col]
        bar = float(metadata[factor_col]["bar"])
        signal_type = str(metadata[factor_col]["signal_type"])

        if signal_type == "state":
            signal = pd.Series(0.0, index=factor.index)
            signal.loc[factor > bar] = 1.0
            signal.loc[factor < bar] = -1.0
        elif signal_type == "event":
            signal = pd.Series(np.nan, index=factor.index, dtype="float64")
            has_event = factor.notna()
            signal.loc[has_event & (factor > bar)] = 1.0
            signal.loc[has_event & (factor < -bar)] = -1.0
            signal.loc[has_event & factor.between(-bar, bar)] = 0.0
        else:
            raise ValueError(f"Unsupported signal_type for {factor_col}: {signal_type!r}")

        signal_ls_df[factor_col] = signal

    return signal_ls_df


def save_factor_outputs(
    mounted_normalized_factor_df: pd.DataFrame,
    signal_ls_df: pd.DataFrame,
    missing_bar_defaults: list[dict[str, object]],
    output_dir: Path | str = output_dir,
    output_prefix: str = "zhao",
    write_empty_missing_bar_file: bool = True,
) -> dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    mounted_factor_path = output_dir / f"{output_prefix}_mounted_normalized_factors.parquet"
    signal_ls_path = output_dir / f"{output_prefix}_signal_ls.parquet"
    mounted_factor_xlsx_path = output_dir / f"{output_prefix}_mounted_normalized_factors.xlsx"
    signal_ls_xlsx_path = output_dir / f"{output_prefix}_signal_ls.xlsx"
    missing_bar_path = output_dir / f"{output_prefix}_missing_bar_defaults.md"

    mounted_normalized_factor_df.to_parquet(mounted_factor_path)
    signal_ls_df.to_parquet(signal_ls_path)
    _write_factor_frame_xlsx(mounted_normalized_factor_df, mounted_factor_xlsx_path)
    _write_factor_frame_xlsx(signal_ls_df, signal_ls_xlsx_path)

    lines = [
        "# Missing bar defaults",
        "",
        "以下因子在 factor_done.json 中 bar 为空，本脚本已按 bar=0 生成 signal_ls。",
        "",
    ]
    if missing_bar_defaults:
        lines.append("| code | signal_type | progress | factor | default_bar |")
        lines.append("| --- | --- | --- | --- | ---: |")
        for item in missing_bar_defaults:
            lines.append(
                "| {code} | {signal_type} | {progress} | {factor} | {default_bar} |".format(
                    code=item.get("code") or item.get("factor_id") or "",
                    signal_type=item["signal_type"],
                    progress=item.get("progress") or "",
                    factor=str(item.get("factor") or "").replace("|", "\\|"),
                    default_bar=item["default_bar"],
                )
            )
    else:
        lines.append("无。")
    if missing_bar_defaults or write_empty_missing_bar_file:
        missing_bar_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    output_paths = {
        "mounted_factor_parquet": mounted_factor_path,
        "signal_ls_parquet": signal_ls_path,
        "mounted_factor_xlsx": mounted_factor_xlsx_path,
        "signal_ls_xlsx": signal_ls_xlsx_path,
    }
    if missing_bar_defaults or write_empty_missing_bar_file:
        output_paths["missing_bar_defaults"] = missing_bar_path
    return output_paths


def _write_factor_frame_xlsx(df: pd.DataFrame, path: Path) -> None:
    out = df.copy()
    out.index = pd.to_datetime(out.index).normalize()
    out.index.name = out.index.name or "date"
    out.to_excel(path)

    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    wb.save(path)
