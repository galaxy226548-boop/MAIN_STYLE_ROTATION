"""Build W003 grouped factor matrix and signal matrix."""

from __future__ import annotations

import json
from collections import OrderedDict
from pathlib import Path

import numpy as np
import pandas as pd


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[2]

FACTOR_GROUP = "W003"
SOURCE_FILE_NAME = "如何从赔率和胜率看成长价值轮动——市场风格轮动系列_mounted_normalized_factors.parquet"
INPUT_PATH = PROJECT_ROOT / "F_grouping" / "input_COMB" / SOURCE_FILE_NAME
OUTPUT_DIR = PROJECT_ROOT / "F_grouping" / "output_COMB"

FACTOR_OUTPUT_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_mounted_normalized_factors.parquet"
SIGNAL_OUTPUT_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_signal_ls.parquet"
FACTOR_XLSX_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_mounted_normalized_factors.xlsx"
SIGNAL_XLSX_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_signal_ls.xlsx"
RECORD_JSON_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_factor_grouping_record.json"


def zscore_series(series: pd.Series, label: str, warnings: list[str]) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    non_null_count = int(numeric.notna().sum())
    if non_null_count < 2:
        warnings.append(f"{label}: non-null sample count is {non_null_count}, cannot z-score with ddof=1.")
        return pd.Series(np.nan, index=series.index, name=series.name, dtype="float64")

    std = numeric.std(skipna=True, ddof=1)
    if pd.isna(std) or np.isclose(std, 0.0):
        warnings.append(f"{label}: standard deviation is {std}, z-score set to NaN.")
        return pd.Series(np.nan, index=series.index, name=series.name, dtype="float64")

    return ((numeric - numeric.mean(skipna=True)) / std).astype("float64")


def group_columns_by_first_letter(columns: pd.Index) -> OrderedDict[str, list[str]]:
    groups: OrderedDict[str, list[str]] = OrderedDict()
    for column in columns:
        col = str(column)
        if not col:
            continue
        groups.setdefault(col[0], []).append(col)
    return groups


def build_factor_matrix(source_df: pd.DataFrame) -> tuple[pd.DataFrame, OrderedDict[str, list[str]], list[str]]:
    warnings: list[str] = []
    source_numeric = source_df.apply(pd.to_numeric, errors="coerce")
    normalized_source = pd.DataFrame(index=source_numeric.index)

    for column in source_numeric.columns:
        normalized_source[column] = zscore_series(source_numeric[column], str(column), warnings)

    groups = group_columns_by_first_letter(source_numeric.columns)
    first_level_raw = pd.DataFrame(index=source_numeric.index)
    first_level_factor = pd.DataFrame(index=source_numeric.index)

    for letter, columns in groups.items():
        factor_col = f"{FACTOR_GROUP}_{letter}"
        first_level_raw[factor_col] = normalized_source[columns].mean(axis=1, skipna=True)
        first_level_factor[factor_col] = zscore_series(first_level_raw[factor_col], factor_col, warnings)

    final_factor = first_level_factor.mean(axis=1, skipna=True).astype("float64")
    final_factor.name = FACTOR_GROUP

    factor_df = first_level_factor.copy()
    factor_df[FACTOR_GROUP] = final_factor
    return factor_df, groups, warnings


def build_signal_ls_matrix(factor_df: pd.DataFrame) -> pd.DataFrame:
    return (factor_df / 3.0).clip(lower=-1.0, upper=1.0).astype("float64")


def write_factor_frame_xlsx(df: pd.DataFrame, path: Path) -> None:
    out = df.copy()
    out.index = pd.to_datetime(out.index).normalize()
    out.index.name = out.index.name or "date"
    out.to_excel(path)

    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    wb.save(path)


def build_record(
    source_df: pd.DataFrame,
    factor_df: pd.DataFrame,
    groups: OrderedDict[str, list[str]],
    warnings: list[str],
) -> dict[str, object]:
    factor_logic = (
        "For each source factor, compute full-sample z-score using "
        "(x - mean(skipna=True)) / std(skipna=True, ddof=1). Then group source factors by "
        "their shared first letter, take the row-wise mean of the normalized source factors, "
        "and z-score that first-level series again with ddof=1."
    )
    signal_logic = "signal_ls = factor_value / 3, clipped to [-1, 1]; NaN remains NaN."
    final_logic = (
        "W003 is the row-wise mean of all first-level W003_{letter} factor values. "
        "It is not z-scored again after averaging."
    )

    first_level_factors = {}
    for letter, columns in groups.items():
        factor_col = f"{FACTOR_GROUP}_{letter}"
        first_level_factors[factor_col] = {
            "shared_prefix": letter,
            "source_columns": columns,
            "factor_value_logic": factor_logic,
            "signal_ls_logic": signal_logic,
            "non_null_count": int(factor_df[factor_col].notna().sum()),
        }

    return {
        "factor_group": FACTOR_GROUP,
        "source_data": {
            "path": str(INPUT_PATH),
            "file_name": INPUT_PATH.name,
            "columns": [str(col) for col in source_df.columns],
            "shape": list(source_df.shape),
            "index_name": source_df.index.name,
            "start_date": str(pd.to_datetime(source_df.index).min().date()),
            "end_date": str(pd.to_datetime(source_df.index).max().date()),
        },
        "first_level_factors": first_level_factors,
        "final_factor": {
            "name": FACTOR_GROUP,
            "source_first_level_factors": [f"{FACTOR_GROUP}_{letter}" for letter in groups],
            "factor_value_logic": final_logic,
            "signal_ls_logic": signal_logic,
            "non_null_count": int(factor_df[FACTOR_GROUP].notna().sum()),
        },
        "outputs": {
            "mounted_normalized_factors_parquet": str(FACTOR_OUTPUT_PATH),
            "mounted_normalized_factors_xlsx": str(FACTOR_XLSX_PATH),
            "signal_ls_parquet": str(SIGNAL_OUTPUT_PATH),
            "signal_ls_xlsx": str(SIGNAL_XLSX_PATH),
            "record_json": str(RECORD_JSON_PATH),
        },
        "warnings": warnings,
    }


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"Input factor matrix not found: {INPUT_PATH}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_df = pd.read_parquet(INPUT_PATH)
    source_df.index = pd.to_datetime(source_df.index)
    source_df = source_df.sort_index()

    factor_df, groups, warnings = build_factor_matrix(source_df)
    signal_ls_df = build_signal_ls_matrix(factor_df)
    record = build_record(source_df, factor_df, groups, warnings)

    factor_df.to_parquet(FACTOR_OUTPUT_PATH)
    signal_ls_df.to_parquet(SIGNAL_OUTPUT_PATH)
    write_factor_frame_xlsx(factor_df, FACTOR_XLSX_PATH)
    write_factor_frame_xlsx(signal_ls_df, SIGNAL_XLSX_PATH)
    RECORD_JSON_PATH.write_text(json.dumps(record, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print("factor_df shape:", factor_df.shape)
    print("signal_ls_df shape:", signal_ls_df.shape)
    print("factor columns:", list(factor_df.columns))
    print("source groups:", {letter: columns for letter, columns in groups.items()})
    print("outputs saved to:", OUTPUT_DIR)
    if warnings:
        print("warnings:")
        for warning in warnings:
            print("-", warning)


if __name__ == "__main__":
    main()
