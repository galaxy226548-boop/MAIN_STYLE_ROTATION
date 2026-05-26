"""计算各大类因子组内相关性并输出热力图。

Usage:
    python F_grouping/scripts/group_correlation.py

Input:
    B_factors/output/grouping_document/factor_{group}_mounted_normalized_factors.parquet
    B_factors/output/grouping_document/factor_{group}_signal_ls.parquet
    F_grouping/reference/usable_factors.xlsx

Output:
    F_grouping/reference/grouping_correlations/
"""

from __future__ import annotations

import argparse
import os
import warnings
from pathlib import Path
from typing import Any

# 避免 Matplotlib 尝试写入不可写的用户级缓存目录。
_MPL_CONFIG_DIR = Path(os.environ.get("MPLCONFIGDIR", "/private/tmp/matplotlib-cache"))
_MPL_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(_MPL_CONFIG_DIR))

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from matplotlib import font_manager
from openpyxl import load_workbook
from scipy.cluster.hierarchy import leaves_list, linkage
from scipy.spatial.distance import squareform
from scipy.stats import ConstantInputWarning, spearmanr


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[2]

INPUT_DIR = PROJECT_ROOT / "B_factors" / "output" / "grouping_document"
METADATA_PATH = PROJECT_ROOT / "F_grouping" / "reference" / "usable_factors.xlsx"
ECONOMIC_SCREEN_PATH = PROJECT_ROOT / "F_grouping" / "reference" / "因子經濟意義篩.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "F_grouping" / "reference" / "grouping_correlations"
CONSTANT_FACTORS_PATH = OUTPUT_DIR / "constant_factors.md"

GROUP_IDS = tuple("CDFGILOPV")
FACTOR_ID_COL = "factor_id"
DEFAULT_DISCREPANCY_THRESHOLD = 0.3
EVENT_NAN_RATE_THRESHOLD = 0.3
DROPNA_WARNING_THRESHOLD = 0.2
DEFAULT_MIN_PERIODS = 10
DEFAULT_MAX_FFILL_DAYS = 5
CHINESE_FONT_PATHS = (
    Path("/System/Library/Fonts/STHeiti Medium.ttc"),
    Path("/System/Library/Fonts/Hiragino Sans GB.ttc"),
    Path("/System/Library/Fonts/Supplemental/Songti.ttc"),
    Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
)
_FONT_CONFIGURED = False


def _format_constant_value(value: object) -> str:
    """将常数值格式化为稳定的 Markdown 文本。"""
    if pd.isna(value):
        return ""
    if isinstance(value, (float, np.floating)):
        return f"{float(value):.12g}"
    return str(value)


def scan_constant_factors(input_dir: Path = INPUT_DIR) -> pd.DataFrame:
    """扫描九大类因子值文件，识别非空值全为同一常数的因子。"""
    records: list[dict[str, object]] = []
    for group_id in GROUP_IDS:
        value_path = input_dir / f"factor_{group_id}_mounted_normalized_factors.parquet"
        if not value_path.exists():
            raise FileNotFoundError(f"Missing value file: {value_path}")

        value_df = pd.read_parquet(value_path)
        for factor_id in value_df.columns.astype(str):
            non_null_series = value_df[factor_id].dropna()
            non_null_count = int(len(non_null_series))
            if non_null_count == 0:
                continue
            if non_null_series.nunique(dropna=True) != 1:
                continue
            records.append(
                {
                    "group_id": group_id,
                    "factor_id": factor_id,
                    "non_null_count": non_null_count,
                    "constant_value": non_null_series.iloc[0],
                }
            )

    if not records:
        return pd.DataFrame(columns=["group_id", "factor_id", "non_null_count", "constant_value"])
    return pd.DataFrame(records).sort_values(["group_id", "factor_id"], kind="mergesort").reset_index(drop=True)


def write_constant_factors_report(constant_df: pd.DataFrame, output_path: Path = CONSTANT_FACTORS_PATH) -> None:
    """将常数因子扫描结果写入 Markdown 报告。"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# 常数因子扫描结果",
        "",
        f"- 扫描目录：`{INPUT_DIR}`",
        "- 扫描范围：`factor_{C,D,F,G,I,L,O,P,V}_mounted_normalized_factors.parquet`",
        "- 判定口径：非空样本数 > 0 且非空唯一值数量 == 1；全 NaN 因子不计入常数因子。",
        "",
        "| group_id | factor_id | non_null_count | constant_value |",
        "| --- | --- | ---: | --- |",
    ]
    if constant_df.empty:
        lines.append("|  |  |  |  |")
    else:
        for row in constant_df.itertuples(index=False):
            lines.append(
                "| "
                f"{row.group_id} | "
                f"{row.factor_id} | "
                f"{row.non_null_count} | "
                f"{_format_constant_value(row.constant_value)} |"
            )
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def remove_factor_rows_from_workbook(path: Path, factor_ids: set[str]) -> dict[str, int]:
    """从工作簿中所有包含 factor_id 表头的 sheet 删除匹配行。"""
    if not factor_ids:
        return {}

    workbook = load_workbook(path)
    removed_by_sheet: dict[str, int] = {}
    workbook_changed = False
    for worksheet in workbook.worksheets:
        header_values = [worksheet.cell(1, column_index).value for column_index in range(1, worksheet.max_column + 1)]
        if FACTOR_ID_COL not in header_values:
            continue

        factor_id_column = header_values.index(FACTOR_ID_COL) + 1
        rows_to_delete: list[int] = []
        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row_index, factor_id_column).value
            if value is not None and str(value).strip() in factor_ids:
                rows_to_delete.append(row_index)

        for row_index in reversed(rows_to_delete):
            worksheet.delete_rows(row_index)

        if rows_to_delete:
            workbook_changed = True
            removed_by_sheet[worksheet.title] = len(rows_to_delete)

    if workbook_changed:
        workbook.save(path)
    return removed_by_sheet


def remove_constant_factors_from_reference_files(constant_df: pd.DataFrame) -> None:
    """根据常数因子名单同步更新可用因子和经济意义筛选表。"""
    factor_ids = set(constant_df["factor_id"].astype(str)) if not constant_df.empty else set()
    if not factor_ids:
        print("constant factors: none found")
        return

    for path in [METADATA_PATH, ECONOMIC_SCREEN_PATH]:
        removed_by_sheet = remove_factor_rows_from_workbook(path, factor_ids)
        if removed_by_sheet:
            print(f"removed constant factors from {path}: {removed_by_sheet}")
        else:
            print(f"removed constant factors from {path}: no matching rows")


def configure_matplotlib_fonts() -> None:
    """配置 Matplotlib 中文字体，避免热力图中文标题缺字。"""
    global _FONT_CONFIGURED
    if _FONT_CONFIGURED:
        return

    for font_path in CHINESE_FONT_PATHS:
        if not font_path.exists():
            continue
        try:
            font_manager.fontManager.addfont(str(font_path))
            font_name = font_manager.FontProperties(fname=str(font_path)).get_name()
        except RuntimeError:
            continue
        plt.rcParams["font.family"] = "sans-serif"
        plt.rcParams["font.sans-serif"] = [font_name, "DejaVu Sans"]
        plt.rcParams["axes.unicode_minus"] = False
        _FONT_CONFIGURED = True
        return

    # 若当前机器没有可用中文字体，也至少保证负号正常显示。
    plt.rcParams["axes.unicode_minus"] = False
    _FONT_CONFIGURED = True


def load_usable_factor_ids(metadata_path: Path = METADATA_PATH) -> dict[str, list[str]]:
    """读取可用因子名录，并按九大类返回 factor_id 列表。"""
    metadata_df = pd.read_excel(metadata_path)
    if FACTOR_ID_COL not in metadata_df.columns:
        raise ValueError(f"Missing column {FACTOR_ID_COL!r} in {metadata_path}")

    factor_ids = metadata_df[FACTOR_ID_COL].dropna().astype(str).str.strip()
    grouped_factor_ids: dict[str, list[str]] = {}
    for group_id in GROUP_IDS:
        # 只处理需求指定的九大类，metadata 中的 W 等其他开头默认忽略。
        group_factor_ids = factor_ids[factor_ids.str.startswith(group_id)].drop_duplicates().tolist()
        grouped_factor_ids[group_id] = group_factor_ids
    return grouped_factor_ids


def read_group_frames(
    group_id: str,
    usable_factor_ids: list[str],
    input_dir: Path = INPUT_DIR,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """读取单个大类的因子值和信号矩阵，并筛选可参与计算的因子列。"""
    value_path = input_dir / f"factor_{group_id}_mounted_normalized_factors.parquet"
    signal_path = input_dir / f"factor_{group_id}_signal_ls.parquet"
    if not value_path.exists():
        raise FileNotFoundError(f"Missing value file: {value_path}")
    if not signal_path.exists():
        raise FileNotFoundError(f"Missing signal file: {signal_path}")

    value_df = pd.read_parquet(value_path)
    signal_df = pd.read_parquet(signal_path)
    value_df.index = pd.to_datetime(value_df.index)
    signal_df.index = pd.to_datetime(signal_df.index)

    # 同时存在于 metadata、因子值文件、信号文件中的列才进入组内相关性计算。
    available_columns = [factor_id for factor_id in usable_factor_ids if factor_id in value_df.columns and factor_id in signal_df.columns]
    if not available_columns:
        raise ValueError(f"No usable factor columns found for group {group_id}")

    missing_columns = sorted(set(usable_factor_ids) - set(available_columns))
    if missing_columns:
        print(f"WARNING: group {group_id} skipped missing factors: {missing_columns}")

    return value_df.loc[:, available_columns], signal_df.loc[:, available_columns], available_columns


def classify_factor_frequency(signal_df: pd.DataFrame) -> dict[str, str]:
    """根据信号 NaN 比例判断因子频率，NaN 率大于 30% 视为 event，否则视为 state。"""
    nan_rates = signal_df.isna().mean()
    return {
        column: "event" if nan_rate > EVENT_NAN_RATE_THRESHOLD else "state"
        for column, nan_rate in nan_rates.items()
    }


def _align_state_frame_to_anchors(df: pd.DataFrame, columns: list[str], anchor_dates: pd.DatetimeIndex) -> pd.DataFrame:
    """将 state 因子对齐到锚点日期，锚点缺失时最多向前取 1 个交易日。"""
    if not columns:
        return pd.DataFrame(index=anchor_dates)

    source_df = df.loc[:, columns].sort_index()
    combined_index = source_df.index.union(anchor_dates).sort_values()
    # 先插入锚点日期，再向前取上一条可用交易日；limit=1 表示最多跨过一个插入的锚点缺口。
    aligned_df = source_df.reindex(combined_index).ffill(limit=1).reindex(anchor_dates)
    return aligned_df


def align_group_frames(
    value_df: pd.DataFrame,
    signal_df: pd.DataFrame,
    frequency_map: dict[str, str],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """按混频规则对齐单个大类的因子值矩阵和信号矩阵。"""
    event_columns = [column for column, frequency in frequency_map.items() if frequency == "event"]
    state_columns = [column for column, frequency in frequency_map.items() if frequency == "state"]
    original_columns = list(value_df.columns)

    if not event_columns:
        # 全部为 state 时不降采样，使用全部可用交易日。
        common_index = value_df.index.intersection(signal_df.index).sort_values()
        return value_df.reindex(common_index), signal_df.reindex(common_index)

    anchor_dates = signal_df.loc[:, event_columns].dropna(how="all").index.sort_values().unique()
    anchor_dates = pd.DatetimeIndex(anchor_dates)

    event_value_df = value_df.loc[:, event_columns].reindex(anchor_dates)
    event_signal_df = signal_df.loc[:, event_columns].reindex(anchor_dates)
    state_value_df = _align_state_frame_to_anchors(value_df, state_columns, anchor_dates)
    state_signal_df = _align_state_frame_to_anchors(signal_df, state_columns, anchor_dates)

    # 重新拼回原始因子顺序，避免对后续聚类前的输入顺序造成额外扰动。
    aligned_value_df = pd.concat([event_value_df, state_value_df], axis=1).reindex(columns=original_columns)
    aligned_signal_df = pd.concat([event_signal_df, state_signal_df], axis=1).reindex(columns=original_columns)
    return aligned_value_df, aligned_signal_df


def filter_sparse_columns(
    df: pd.DataFrame,
    group_id: str,
    frame_name: str,
    min_periods: int = DEFAULT_MIN_PERIODS,
) -> tuple[pd.DataFrame, list[str]]:
    """排除全空或样本数不足的因子列，并打印中文警告。"""
    kept_columns: list[str] = []
    excluded_columns: list[str] = []
    non_null_counts = df.notna().sum()

    for column, count in non_null_counts.items():
        if count == 0:
            print(f"WARNING: group {group_id} {frame_name} factor {column} excluded: 全列为 NaN")
            excluded_columns.append(str(column))
        elif count < min_periods:
            print(
                f"WARNING: group {group_id} {frame_name} factor {column} excluded: "
                f"非空样本数 {count} < min_periods {min_periods}"
            )
            excluded_columns.append(str(column))
        else:
            kept_columns.append(str(column))

    return df.loc[:, kept_columns], excluded_columns


def _choose_anchor_factor(
    factor_a: str,
    factor_b: str,
    non_null_counts: pd.Series,
) -> tuple[str, str]:
    """为单个因子对选择锚点因子和填充因子。"""
    count_a = int(non_null_counts[factor_a])
    count_b = int(non_null_counts[factor_b])
    if count_a < count_b:
        return factor_a, factor_b
    if count_b < count_a:
        return factor_b, factor_a
    return (factor_a, factor_b) if factor_a <= factor_b else (factor_b, factor_a)


def _ffill_to_anchor_dates(
    fill_series: pd.Series,
    anchor_dates: pd.DatetimeIndex,
    max_ffill_days: int,
) -> pd.Series:
    """将填充因子单向 ffill 到锚点日期，并限制自然日追溯窗口。"""
    fill_non_null = fill_series.dropna().sort_index()
    if fill_non_null.empty or len(anchor_dates) == 0:
        return pd.Series(index=anchor_dates, dtype=float)

    combined_index = fill_non_null.index.union(anchor_dates).sort_values()
    value_series = fill_non_null.reindex(combined_index)
    source_dates = pd.Series(pd.NaT, index=combined_index, dtype="datetime64[ns]")
    source_dates.loc[fill_non_null.index] = fill_non_null.index

    # ffill 后用 source_date 判断追溯自然日是否超过阈值。
    filled_values = value_series.ffill().reindex(anchor_dates)
    filled_source_dates = source_dates.ffill().reindex(anchor_dates)
    age_days = (pd.Series(anchor_dates, index=anchor_dates) - filled_source_dates).dt.days
    return filled_values.where(age_days <= max_ffill_days)


def _pairwise_aligned_series(
    df: pd.DataFrame,
    factor_a: str,
    factor_b: str,
    non_null_counts: pd.Series,
    max_ffill_days: int,
) -> tuple[pd.Series, pd.Series]:
    """按低频锚点和单向 ffill 规则对齐单个因子对。"""
    anchor_factor, fill_factor = _choose_anchor_factor(factor_a, factor_b, non_null_counts)
    anchor_series = df[anchor_factor].dropna().sort_index()
    anchor_dates = pd.DatetimeIndex(anchor_series.index)
    fill_series = _ffill_to_anchor_dates(df[fill_factor], anchor_dates, max_ffill_days)

    if factor_a == anchor_factor:
        return anchor_series, fill_series
    return fill_series, anchor_series


def compute_pairwise_spearman(
    df: pd.DataFrame,
    min_periods: int = DEFAULT_MIN_PERIODS,
    max_ffill_days: int = DEFAULT_MAX_FFILL_DAYS,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    对 df 中每对列按低频锚点和单向 ffill 规则计算 Spearman 相关系数。

    返回：
        corr_matrix: 相关性矩阵，shape (n, n)
        n_matrix: 每对因子的有效样本数矩阵，shape (n, n)
    """
    columns = list(df.columns)
    corr_matrix = pd.DataFrame(np.nan, index=columns, columns=columns, dtype=float)
    n_matrix = pd.DataFrame(0, index=columns, columns=columns, dtype=int)
    non_null_counts = df.notna().sum()

    for left_index, factor_a in enumerate(columns):
        for factor_b in columns[left_index:]:
            if factor_a == factor_b:
                aligned_a = df[factor_a].dropna().sort_index()
                aligned_b = aligned_a.copy()
            else:
                aligned_a, aligned_b = _pairwise_aligned_series(
                    df,
                    factor_a,
                    factor_b,
                    non_null_counts,
                    max_ffill_days,
                )

            pair_df = pd.concat([aligned_a.rename("left"), aligned_b.rename("right")], axis=1).dropna(how="any")
            valid_n = int(len(pair_df))
            n_matrix.loc[factor_a, factor_b] = valid_n
            n_matrix.loc[factor_b, factor_a] = valid_n
            if valid_n < min_periods:
                continue

            with warnings.catch_warnings():
                warnings.simplefilter("ignore", ConstantInputWarning)
                corr_value = spearmanr(pair_df["left"], pair_df["right"], nan_policy="omit").statistic
            corr_matrix.loc[factor_a, factor_b] = corr_value
            corr_matrix.loc[factor_b, factor_a] = corr_value

    return corr_matrix, n_matrix


def cluster_factor_order(value_corr: pd.DataFrame) -> tuple[list[str], np.ndarray | None]:
    """基于因子值相关性矩阵做层次聚类，并返回因子排序和 linkage 矩阵。"""
    columns = list(value_corr.columns)
    n_factors = len(columns)
    if n_factors < 2:
        return columns, None

    # 聚类距离使用 1 - |corr|；无法计算的相关性按 0 处理，避免 linkage 因 NaN 失败。
    corr_abs_array = value_corr.abs().reindex(index=columns, columns=columns).fillna(0.0).to_numpy(copy=True)
    np.fill_diagonal(corr_abs_array, 1.0)
    distance_array = 1.0 - corr_abs_array
    np.fill_diagonal(distance_array, 0.0)
    condensed_distance = squareform(distance_array, checks=False)

    linkage_matrix = linkage(condensed_distance, method="average")
    ordered_columns = [columns[index] for index in leaves_list(linkage_matrix)]
    return ordered_columns, linkage_matrix


def suggest_cluster_count(linkage_matrix: np.ndarray | None, n_factors: int) -> int:
    """用层次聚类距离跳变法推断建议子类数量。"""
    if n_factors < 3 or linkage_matrix is None or len(linkage_matrix) < 2:
        return 1

    distances = linkage_matrix[:, 2]
    if not np.isfinite(distances).all():
        return 1

    jumps = np.diff(distances)
    if len(jumps) == 0 or not np.isfinite(jumps).all():
        return 1

    jump_index = int(np.argmax(jumps))
    suggested_k = n_factors - (jump_index + 1)
    return int(min(max(suggested_k, 1), n_factors))


def mean_upper_triangle(corr_df: pd.DataFrame) -> float:
    """计算相关性矩阵上三角不含对角线元素均值。"""
    n_factors = len(corr_df.columns)
    if n_factors < 2:
        return float("nan")

    upper_rows, upper_cols = np.triu_indices(n_factors, k=1)
    values = corr_df.to_numpy()[upper_rows, upper_cols]
    if np.isnan(values).all():
        return float("nan")
    return float(np.nanmean(values))


def upper_triangle_values(df: pd.DataFrame) -> np.ndarray:
    """提取矩阵上三角不含对角线的元素。"""
    n_factors = len(df.columns)
    if n_factors < 2:
        return np.array([], dtype=float)
    upper_rows, upper_cols = np.triu_indices(n_factors, k=1)
    return df.to_numpy(dtype=float)[upper_rows, upper_cols]


def pairwise_n_description(n_matrix: pd.DataFrame) -> str:
    """生成热力图副标题使用的有效样本数范围描述。"""
    values = upper_triangle_values(n_matrix)
    values = values[np.isfinite(values)]
    if len(values) == 0:
        return "有效样本数范围：NA ～ NA 行（中位数 NA 行）"
    min_n = int(np.nanmin(values))
    max_n = int(np.nanmax(values))
    median_n = int(np.nanmedian(values))
    return f"有效样本数范围：{min_n} ～ {max_n} 行（中位数 {median_n} 行）"


def pairwise_n_summary(n_matrix: pd.DataFrame) -> tuple[float, float]:
    """返回上三角有效样本数的最小值和中位数。"""
    values = upper_triangle_values(n_matrix)
    values = values[np.isfinite(values)]
    if len(values) == 0:
        return float("nan"), float("nan")
    return float(np.nanmin(values)), float(np.nanmedian(values))


def count_insufficient_pairs(
    value_corr: pd.DataFrame,
    signal_corr: pd.DataFrame,
    value_n: pd.DataFrame,
    signal_n: pd.DataFrame,
    min_periods: int = DEFAULT_MIN_PERIODS,
) -> int:
    """统计因样本不足导致任一相关性矩阵为 NaN 的因子对数量。"""
    columns = list(value_corr.columns)
    n_nan_pairs = 0
    for left_index, factor_a in enumerate(columns):
        for factor_b in columns[left_index + 1:]:
            value_is_nan = pd.isna(value_corr.loc[factor_a, factor_b])
            signal_is_nan = pd.isna(signal_corr.loc[factor_a, factor_b])
            if not (value_is_nan or signal_is_nan):
                continue
            value_count = value_n.loc[factor_a, factor_b]
            signal_count = signal_n.loc[factor_a, factor_b]
            value_insufficient = pd.isna(value_count) or value_count < min_periods
            signal_insufficient = pd.isna(signal_count) or signal_count < min_periods
            if value_insufficient or signal_insufficient:
                n_nan_pairs += 1
    return n_nan_pairs


def plot_group_heatmaps(
    group_id: str,
    value_corr: pd.DataFrame,
    signal_corr: pd.DataFrame,
    value_n: pd.DataFrame,
    signal_n: pd.DataFrame,
    ordered_columns: list[str],
    value_excluded_columns: list[str],
    signal_excluded_columns: list[str],
    output_dir: Path = OUTPUT_DIR,
) -> Path:
    """绘制单个大类的因子值和信号 Spearman 相关性热力图。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    n_factors = len(ordered_columns)
    value_plot_df = value_corr.reindex(index=ordered_columns, columns=ordered_columns)
    signal_plot_df = signal_corr.reindex(index=ordered_columns, columns=ordered_columns)
    value_n_plot_df = value_n.reindex(index=ordered_columns, columns=ordered_columns)
    signal_n_plot_df = signal_n.reindex(index=ordered_columns, columns=ordered_columns)

    annot = n_factors <= 15
    figure_width = min(max(14.0, n_factors * 0.42), 42.0)
    figure_height = min(max(7.0, n_factors * 0.26), 24.0)
    fig, axes = plt.subplots(1, 2, figsize=(figure_width, figure_height), constrained_layout=True)
    cmap = plt.get_cmap("RdYlGn").copy()
    cmap.set_bad("#eeeeee")

    common_kwargs: dict[str, Any] = {
        "cmap": cmap,
        "vmin": -1,
        "vmax": 1,
        "center": 0,
        "square": False,
        "annot": annot,
        "fmt": ".2f",
        "linewidths": 0.0,
        "cbar_kws": {"shrink": 0.75},
    }
    sns.heatmap(value_plot_df, ax=axes[0], **common_kwargs)
    axes[0].set_title(
        f"{group_id}类 — 因子值Spearman相关性（共{n_factors}个因子）\n"
        f"{pairwise_n_description(value_n_plot_df)}"
    )
    sns.heatmap(signal_plot_df, ax=axes[1], **common_kwargs)
    axes[1].set_title(
        f"{group_id}类 — 信号Spearman相关性（共{n_factors}个因子）\n"
        f"{pairwise_n_description(signal_n_plot_df)}"
    )

    excluded_by_axis = [set(value_excluded_columns), set(signal_excluded_columns)]

    for axis, excluded_columns in zip(axes, excluded_by_axis):
        axis.tick_params(axis="x", labelrotation=90, labelsize=7 if n_factors > 40 else 8)
        axis.tick_params(axis="y", labelrotation=0, labelsize=7 if n_factors > 40 else 8)
        for label in [*axis.get_xticklabels(), *axis.get_yticklabels()]:
            if label.get_text() in excluded_columns:
                label.set_color("#888888")

    output_path = output_dir / f"group_{group_id}_correlation_heatmaps.png"
    fig.savefig(output_path, dpi=200)
    plt.close(fig)
    return output_path


def build_discrepancy_table(
    group_id: str,
    value_corr: pd.DataFrame,
    signal_corr: pd.DataFrame,
    discrepancy_threshold: float = DEFAULT_DISCREPANCY_THRESHOLD,
) -> pd.DataFrame:
    """构造因子值相关性与信号相关性差异较大的因子对表。"""
    records: list[dict[str, object]] = []
    columns = list(value_corr.columns)
    for left_index, factor_a in enumerate(columns):
        for factor_b in columns[left_index + 1:]:
            value = value_corr.loc[factor_a, factor_b]
            signal = signal_corr.loc[factor_a, factor_b]
            if pd.isna(value) or pd.isna(signal):
                continue
            discrepancy = abs(float(value) - float(signal))
            if discrepancy > discrepancy_threshold:
                records.append(
                    {
                        "group_id": group_id,
                        "factor_a": factor_a,
                        "factor_b": factor_b,
                        "value_corr": float(value),
                        "signal_corr": float(signal),
                        "discrepancy": discrepancy,
                    }
                )

    output_df = pd.DataFrame(
        records,
        columns=["group_id", "factor_a", "factor_b", "value_corr", "signal_corr", "discrepancy"],
    )
    if not output_df.empty:
        output_df = output_df.sort_values("discrepancy", ascending=False, kind="mergesort").reset_index(drop=True)
    return output_df


def save_group_matrices(
    group_id: str,
    value_corr: pd.DataFrame,
    signal_corr: pd.DataFrame,
    value_n: pd.DataFrame,
    signal_n: pd.DataFrame,
    ordered_columns: list[str],
    output_dir: Path = OUTPUT_DIR,
) -> Path:
    """保存单个大类的因子值和信号相关性矩阵。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"group_{group_id}_correlation_matrices.xlsx"
    with pd.ExcelWriter(output_path) as writer:
        value_corr.reindex(index=ordered_columns, columns=ordered_columns).to_excel(writer, sheet_name="value_corr")
        signal_corr.reindex(index=ordered_columns, columns=ordered_columns).to_excel(writer, sheet_name="signal_corr")
        value_n.reindex(index=ordered_columns, columns=ordered_columns).to_excel(writer, sheet_name="value_n")
        signal_n.reindex(index=ordered_columns, columns=ordered_columns).to_excel(writer, sheet_name="signal_n")
    return output_path


def analyze_group(
    group_id: str,
    usable_factor_ids: list[str],
    discrepancy_threshold: float = DEFAULT_DISCREPANCY_THRESHOLD,
    output_dir: Path = OUTPUT_DIR,
    save_matrices: bool = True,
) -> dict[str, object]:
    """分析单个大类，并返回汇总统计、差异表和输出文件路径。"""
    value_df, signal_df, factor_columns = read_group_frames(group_id, usable_factor_ids)
    frequency_map = classify_factor_frequency(signal_df)
    aligned_value_df, aligned_signal_df = align_group_frames(value_df, signal_df, frequency_map)

    value_input_df, value_excluded_columns = filter_sparse_columns(aligned_value_df, group_id, "value")
    signal_input_df, signal_excluded_columns = filter_sparse_columns(aligned_signal_df, group_id, "signal")

    value_corr, value_n = compute_pairwise_spearman(value_input_df)
    signal_corr, signal_n = compute_pairwise_spearman(signal_input_df)

    ordered_valid_columns, linkage_matrix = cluster_factor_order(value_corr)
    ordered_columns = [
        *ordered_valid_columns,
        *[column for column in factor_columns if column not in ordered_valid_columns],
    ]
    suggested_k = suggest_cluster_count(linkage_matrix, len(ordered_valid_columns))

    value_corr = value_corr.reindex(index=factor_columns, columns=factor_columns)
    signal_corr = signal_corr.reindex(index=factor_columns, columns=factor_columns)
    value_n = value_n.reindex(index=factor_columns, columns=factor_columns)
    signal_n = signal_n.reindex(index=factor_columns, columns=factor_columns)
    combined_n = value_n.combine(signal_n, np.minimum)

    heatmap_path = plot_group_heatmaps(
        group_id,
        value_corr,
        signal_corr,
        value_n,
        signal_n,
        ordered_columns,
        value_excluded_columns,
        signal_excluded_columns,
        output_dir,
    )
    matrices_path = (
        save_group_matrices(group_id, value_corr, signal_corr, value_n, signal_n, ordered_columns, output_dir)
        if save_matrices
        else None
    )
    discrepancy_df = build_discrepancy_table(group_id, value_corr, signal_corr, discrepancy_threshold)

    n_factors = len(factor_columns)
    min_pairwise_n, median_pairwise_n = pairwise_n_summary(combined_n)
    summary = {
        "group_id": group_id,
        "mean_value_corr": mean_upper_triangle(value_corr),
        "mean_signal_corr": mean_upper_triangle(signal_corr),
        "n_factors": n_factors,
        "n_pairs": n_factors * (n_factors - 1) // 2,
        "suggested_k": suggested_k,
        "min_pairwise_n": min_pairwise_n,
        "median_pairwise_n": median_pairwise_n,
        "n_nan_pairs": count_insufficient_pairs(value_corr, signal_corr, value_n, signal_n),
        "n_state": sum(frequency == "state" for frequency in frequency_map.values()),
        "n_event": sum(frequency == "event" for frequency in frequency_map.values()),
        "heatmap_path": str(heatmap_path),
        "matrices_path": str(matrices_path) if matrices_path else "",
    }
    print(summary)
    return {
        "summary": summary,
        "discrepancy_df": discrepancy_df,
        "value_corr": value_corr,
        "signal_corr": signal_corr,
        "value_n": value_n,
        "signal_n": signal_n,
        "ordered_columns": ordered_columns,
    }


def run_all_groups(
    discrepancy_threshold: float = DEFAULT_DISCREPANCY_THRESHOLD,
    output_dir: Path = OUTPUT_DIR,
    save_matrices: bool = True,
) -> dict[str, dict[str, object]]:
    """运行九大类组内相关性分析，并保存汇总统计和差异表。"""
    configure_matplotlib_fonts()
    output_dir.mkdir(parents=True, exist_ok=True)
    constant_df = scan_constant_factors()
    write_constant_factors_report(constant_df, output_dir / CONSTANT_FACTORS_PATH.name)
    remove_constant_factors_from_reference_files(constant_df)
    usable_factor_ids_by_group = load_usable_factor_ids()
    results: dict[str, dict[str, object]] = {}

    for group_id in GROUP_IDS:
        results[group_id] = analyze_group(
            group_id,
            usable_factor_ids_by_group[group_id],
            discrepancy_threshold=discrepancy_threshold,
            output_dir=output_dir,
            save_matrices=save_matrices,
        )

    summary_df = pd.DataFrame([result["summary"] for result in results.values()])
    summary_path = output_dir / "group_correlation_summary.xlsx"
    summary_df.to_excel(summary_path, index=False)
    print("cross-group summary:")
    print(summary_df.to_string(index=False))

    discrepancy_frames = [result["discrepancy_df"] for result in results.values()]
    discrepancy_df = pd.concat(discrepancy_frames, ignore_index=True)
    if not discrepancy_df.empty:
        discrepancy_df = discrepancy_df.sort_values("discrepancy", ascending=False, kind="mergesort").reset_index(drop=True)
    discrepancy_path = output_dir / "group_correlation_discrepancies.xlsx"
    discrepancy_df.to_excel(discrepancy_path, index=False)

    print(f"summary saved to: {summary_path}")
    print(f"discrepancies saved to: {discrepancy_path}")
    return results


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""
    parser = argparse.ArgumentParser(description="计算各大类因子组内相关性并输出热力图。")
    parser.add_argument(
        "--discrepancy-threshold",
        type=float,
        default=DEFAULT_DISCREPANCY_THRESHOLD,
        help="因子值相关性与信号相关性差异阈值，默认 0.3。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIR,
        help="输出目录，默认 F_grouping/reference/grouping_correlations。",
    )
    parser.add_argument(
        "--no-save-matrices",
        action="store_true",
        help="不保存每组相关性矩阵 Excel。",
    )
    return parser.parse_args()


def main() -> None:
    """脚本入口。"""
    args = parse_args()
    run_all_groups(
        discrepancy_threshold=args.discrepancy_threshold,
        output_dir=args.output_dir,
        save_matrices=not args.no_save_matrices,
    )


if __name__ == "__main__":
    main()
