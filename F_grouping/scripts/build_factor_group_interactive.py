"""Interactive builder for grouped COMB factor matrix and signal matrix.

Usage:
    python F_grouping/scripts/build_factor_group_interactive.py

Input:
    The script scans:
        F_grouping/input_COMB/*_mounted_normalized_factors.parquet

    In terminal, enter:
        1. factor group name, for example W003
        2. source factor names, for example L002,L003,M001,V079

Output:
    F_grouping/output_COMB/{FACTOR_GROUP}_mounted_normalized_factors.parquet
    F_grouping/output_COMB/{FACTOR_GROUP}_factor_signal_ls.parquet
    F_grouping/output_COMB/{FACTOR_GROUP}_mounted_normalized_factors.xlsx
    F_grouping/output_COMB/{FACTOR_GROUP}_factor_signal_ls.xlsx
    F_grouping/output_COMB/{FACTOR_GROUP}_factor_grouping_record.json
"""

from __future__ import annotations

import ast
import json
import sys
from collections import OrderedDict
from pathlib import Path

import numpy as np
import pandas as pd


# ========== 1. Path config ==========

SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[2]

INPUT_DIR = PROJECT_ROOT / "F_grouping" / "input_COMB"
OUTPUT_DIR = PROJECT_ROOT / "F_grouping" / "output_COMB"
REFERENCE_DIR = PROJECT_ROOT / "F_grouping" / "reference"
INPUT_INVENTORY_PATH = REFERENCE_DIR / "input_comb_inventory.xlsx"

INPUT_FILE_SUFFIX = "_mounted_normalized_factors.parquet"


# ========== 2. Basic utilities ==========

def parse_factor_list(text: str) -> list[str]:
    """Parse terminal input into a clean factor-name list.

    Supported formats:
        L002,L003,M001
        L002 L003 M001
        ["L002", "L003", "M001"]
        'L002', 'L003', 'M001'
    """
    text = text.strip()
    if not text:
        raise ValueError("Factor list is empty.")

    # Try Python-list style first.
    if text.startswith("[") and text.endswith("]"):
        parsed = ast.literal_eval(text)
        if not isinstance(parsed, list):
            raise ValueError("Input looks like a Python object, but it is not a list.")
        factors = [str(x).strip() for x in parsed]
    else:
        # Support comma-separated or whitespace-separated input.
        text = text.replace("，", ",")
        if "," in text:
            factors = [x.strip().strip("'").strip('"') for x in text.split(",")]
        else:
            factors = [x.strip().strip("'").strip('"') for x in text.split()]

    factors = [x for x in factors if x]
    if not factors:
        raise ValueError("No valid factor name is parsed.")

    # Deduplicate while preserving order.
    deduped = list(OrderedDict.fromkeys(factors))
    return deduped


def zscore_series(series: pd.Series, label: str, warnings: list[str]) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    non_null_count = int(numeric.notna().sum())

    if non_null_count < 2:
        warnings.append(
            f"{label}: non-null sample count is {non_null_count}, "
            "cannot z-score with ddof=1."
        )
        return pd.Series(np.nan, index=series.index, name=series.name, dtype="float64")

    std = numeric.std(skipna=True, ddof=1)
    if pd.isna(std) or np.isclose(std, 0.0):
        warnings.append(f"{label}: standard deviation is {std}, z-score set to NaN.")
        return pd.Series(np.nan, index=series.index, name=series.name, dtype="float64")

    return ((numeric - numeric.mean(skipna=True)) / std).astype("float64")


def group_columns_by_first_letter(columns: pd.Index) -> OrderedDict[str, list[str]]:
    groups: OrderedDict[str, list[str]] = OrderedDict()

    for column in columns:
        col = str(column).strip()
        if not col:
            continue
        groups.setdefault(col[0], []).append(col)

    return groups


# ========== 3. Scan and collect source factor columns ==========

def list_input_files(input_dir: Path) -> list[Path]:
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

    files = sorted(input_dir.glob(f"*{INPUT_FILE_SUFFIX}"))
    if not files:
        raise FileNotFoundError(
            f"No input parquet files found under {input_dir} "
            f"with suffix {INPUT_FILE_SUFFIX!r}."
        )

    return files


def parse_inventory_columns(value: object) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value]

    if pd.isna(value):
        return []

    text = str(value).strip()
    if not text:
        return []

    try:
        parsed = ast.literal_eval(text)
    except (ValueError, SyntaxError):
        return []

    if not isinstance(parsed, list):
        return []

    return [str(item) for item in parsed]


def load_inventory_factor_index(
    inventory_path: Path,
    input_dir: Path,
) -> tuple[dict[str, list[Path]], dict[str, object], list[str]]:
    """Load input_comb_inventory.xlsx into a factor -> mounted parquet path index."""
    warnings: list[str] = []
    if not inventory_path.exists():
        warnings.append(f"Inventory file not found; falling back to parquet scan: {inventory_path}")
        return {}, {"path": str(inventory_path), "status": "missing"}, warnings

    try:
        inventory_df = pd.read_excel(inventory_path)
    except Exception as exc:  # noqa: BLE001
        warnings.append(f"Failed to read inventory; falling back to parquet scan: {inventory_path}: {exc}")
        return {}, {"path": str(inventory_path), "status": "read_failed", "error": str(exc)}, warnings

    factor_to_paths: dict[str, list[Path]] = {}
    rows: list[dict[str, object]] = []

    for _, row in inventory_df.iterrows():
        file_name = str(row.get("file_name", "")).strip()
        read_status = str(row.get("read_status", "")).strip().lower()
        if not file_name.endswith(INPUT_FILE_SUFFIX):
            continue
        if read_status and read_status != "success":
            continue

        relative_path = str(row.get("relative_path", "")).strip() or file_name
        path = input_dir / relative_path
        columns = parse_inventory_columns(row.get("first_raw_values"))
        rows.append(
            {
                "file_name": file_name,
                "path": str(path),
                "columns": columns,
                "read_status": row.get("read_status"),
            }
        )
        for column in columns:
            factor_to_paths.setdefault(column, []).append(path)

    inventory_record = {
        "path": str(inventory_path),
        "status": "loaded",
        "mounted_file_count": len(rows),
        "mounted_files": rows,
    }
    return factor_to_paths, inventory_record, warnings


def read_input_file(path: Path) -> pd.DataFrame:
    df = pd.read_parquet(path)

    if df.empty:
        return df

    df.index = pd.to_datetime(df.index)
    df = df.sort_index()

    # Make all column names string to avoid matching failure.
    df.columns = [str(col) for col in df.columns]
    return df


def scan_factor_locations(
    factor_names: list[str],
    input_files: list[Path],
) -> tuple[dict[str, list[Path]], list[dict[str, object]]]:
    factor_to_paths: dict[str, list[Path]] = {name: [] for name in factor_names}
    scanned_files: list[dict[str, object]] = []

    for path in input_files:
        df = read_input_file(path)
        columns = list(df.columns)
        scanned_files.append(
            {
                "file_name": path.name,
                "path": str(path),
                "shape": list(df.shape),
                "columns": columns,
            }
        )
        for factor in factor_names:
            if factor in columns:
                factor_to_paths[factor].append(path)

    return factor_to_paths, scanned_files


def summarize_input_files_from_inventory(
    inventory_record: dict[str, object],
    input_files: list[Path],
) -> list[dict[str, object]]:
    mounted_files = inventory_record.get("mounted_files")
    if isinstance(mounted_files, list) and mounted_files:
        return mounted_files
    return [{"file_name": path.name, "path": str(path)} for path in input_files]


def collect_selected_factors(
    factor_names: list[str],
    input_files: list[Path],
) -> tuple[pd.DataFrame, dict[str, object], list[str]]:
    """Find selected factor columns from all input files.

    Rules:
        - Search all parquet files under input_COMB.
        - If a factor appears in exactly one file, use it.
        - If a factor appears in multiple files, raise ValueError to avoid ambiguous source data.
        - If a factor is missing, raise ValueError.
        - Output columns follow the user's input order.
    """
    warnings: list[str] = []
    inventory_index, inventory_record, inventory_warnings = load_inventory_factor_index(
        INPUT_INVENTORY_PATH,
        INPUT_DIR,
    )
    warnings.extend(inventory_warnings)

    factor_to_paths: dict[str, list[Path]] = {name: list(inventory_index.get(name, [])) for name in factor_names}
    missing_from_inventory = [factor for factor, paths in factor_to_paths.items() if not paths]
    scanned_files: list[dict[str, object]] = []

    scan_method = "inventory"
    if missing_from_inventory:
        scan_method = "inventory_plus_parquet_fallback" if inventory_index else "parquet_scan"
        fallback_locations, scanned_files = scan_factor_locations(missing_from_inventory, input_files)
        for factor, paths in fallback_locations.items():
            factor_to_paths[factor] = paths

    missing = [factor for factor, paths in factor_to_paths.items() if not paths]
    if missing:
        available_columns: list[str] = []
        for item in summarize_input_files_from_inventory(inventory_record, input_files):
            available_columns.extend(item.get("columns", []))
        for item in scanned_files:
            available_columns.extend(item.get("columns", []))

        available_preview = sorted(set(available_columns))[:80]
        raise ValueError(
            "Some requested factors are not found in input_COMB parquet files:\n"
            f"missing factors: {missing}\n\n"
            "Preview of available columns, first 80 sorted unique names:\n"
            f"{available_preview}"
        )

    duplicated_sources = {
        factor: sorted({path.name for path in paths})
        for factor, paths in factor_to_paths.items()
        if len({path.resolve() for path in paths}) > 1
    }
    if duplicated_sources:
        raise ValueError(
            "Some requested factors are found in multiple mounted factor files. "
            "Please adjust input_COMB or choose an unambiguous factor list.\n"
            f"duplicated factor sources: {duplicated_sources}"
        )

    needed_paths = sorted({paths[0] for paths in factor_to_paths.values()})
    data_by_path = {path: read_input_file(path) for path in needed_paths}

    selected_series: list[pd.Series] = []
    factor_source_map: dict[str, dict[str, object]] = {}

    for factor in factor_names:
        source_path = factor_to_paths[factor][0]
        df = data_by_path[source_path]
        if factor not in df.columns:
            raise ValueError(
                f"{factor} was resolved to {source_path.name}, but the column is absent when reading the parquet."
            )
        series = df[factor]
        series = pd.to_numeric(series, errors="coerce")
        series.name = factor
        selected_series.append(series)

        factor_source_map[factor] = {
            "source_file": source_path.name,
            "source_path": str(source_path),
            "non_null_count": int(series.notna().sum()),
            "start_date": (
                str(pd.to_datetime(series.dropna().index).min().date())
                if series.notna().any()
                else None
            ),
            "end_date": (
                str(pd.to_datetime(series.dropna().index).max().date())
                if series.notna().any()
                else None
            ),
        }

    source_df = pd.concat(selected_series, axis=1).sort_index()
    source_df = source_df[factor_names]

    scan_record = {
        "input_dir": str(INPUT_DIR),
        "input_file_suffix": INPUT_FILE_SUFFIX,
        "inventory": inventory_record,
        "scan_method": scan_method,
        "scanned_file_count": len(input_files),
        "input_files": summarize_input_files_from_inventory(inventory_record, input_files),
        "fallback_scanned_files": scanned_files,
        "factor_source_map": factor_source_map,
    }

    return source_df, scan_record, warnings


# ========== 4. Build factor and signal matrix ==========

def build_factor_matrix(
    source_df: pd.DataFrame,
    factor_group: str,
) -> tuple[pd.DataFrame, OrderedDict[str, list[str]], list[str]]:
    warnings: list[str] = []

    source_numeric = source_df.apply(pd.to_numeric, errors="coerce")
    normalized_source = pd.DataFrame(index=source_numeric.index)

    for column in source_numeric.columns:
        normalized_source[column] = zscore_series(source_numeric[column], str(column), warnings)

    groups = group_columns_by_first_letter(source_numeric.columns)

    first_level_raw = pd.DataFrame(index=source_numeric.index)
    first_level_factor = pd.DataFrame(index=source_numeric.index)

    for letter, columns in groups.items():
        factor_col = f"{factor_group}_{letter}"

        # Same as old logic:
        # 1. row-wise mean of normalized bottom-level source factors
        # 2. z-score the first-level raw series again
        first_level_raw[factor_col] = normalized_source[columns].mean(axis=1, skipna=True)
        first_level_factor[factor_col] = zscore_series(
            first_level_raw[factor_col],
            factor_col,
            warnings,
        )

    # Same as old logic:
    # final factor is row-wise mean of all first-level factor values.
    # It is not z-scored again.
    final_factor = first_level_factor.mean(axis=1, skipna=True).astype("float64")
    final_factor.name = factor_group

    factor_df = first_level_factor.copy()
    factor_df[factor_group] = final_factor

    return factor_df, groups, warnings


def build_signal_ls_matrix(factor_df: pd.DataFrame) -> pd.DataFrame:
    return (factor_df / 3.0).clip(lower=-1.0, upper=1.0).astype("float64")


# ========== 5. Output writers ==========

def write_factor_frame_xlsx(df: pd.DataFrame, path: Path) -> None:
    out = df.copy()
    out.index = pd.to_datetime(out.index).normalize()
    out.index.name = out.index.name or "date"
    out.to_excel(path)

    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active

    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"

    wb.save(path)


def build_output_paths(factor_group: str) -> dict[str, Path]:
    return {
        "factor_parquet": OUTPUT_DIR / f"{factor_group}_mounted_normalized_factors.parquet",
        "signal_parquet": OUTPUT_DIR / f"{factor_group}_factor_signal_ls.parquet",
        "factor_xlsx": OUTPUT_DIR / f"{factor_group}_mounted_normalized_factors.xlsx",
        "signal_xlsx": OUTPUT_DIR / f"{factor_group}_factor_signal_ls.xlsx",
        "record_json": OUTPUT_DIR / f"{factor_group}_factor_grouping_record.json",
    }


def find_existing_outputs(output_paths: dict[str, Path]) -> list[Path]:
    return sorted(path for path in output_paths.values() if path.exists() and path.is_file())


def confirm_overwrite_if_needed(output_paths: dict[str, Path]) -> None:
    existing = find_existing_outputs(output_paths)
    if not existing:
        return

    print("\n========== 疑似重复编号 ==========")
    print(f"output_COMB 中已存在 {len(existing)} 个同名 factor-group 输出文件，继续将覆盖：")
    for path in existing:
        print("-", path.name)
    answer = input("确认继续吗？请输入 y/yes 确认：").strip().lower()
    if answer not in {"y", "yes"}:
        print("已取消，未写入任何输出文件。")
        sys.exit(0)


def build_record(
    factor_group: str,
    requested_factor_names: list[str],
    source_df: pd.DataFrame,
    factor_df: pd.DataFrame,
    groups: OrderedDict[str, list[str]],
    scan_record: dict[str, object],
    warnings: list[str],
    output_paths: dict[str, Path],
) -> dict[str, object]:
    factor_logic = (
        "For each selected source factor, compute full-sample z-score using "
        "(x - mean(skipna=True)) / std(skipna=True, ddof=1). Then group selected source "
        "factors by their shared first letter, take the row-wise mean of the normalized "
        "source factors, and z-score that first-level series again with ddof=1."
    )

    signal_logic = "signal_ls = factor_value / 3, clipped to [-1, 1]; NaN remains NaN."

    final_logic = (
        f"{factor_group} is the row-wise mean of all first-level "
        f"{factor_group}" + "_{letter} factor values. "
        "It is not z-scored again after averaging."
    )

    first_level_factors = {}
    for letter, columns in groups.items():
        factor_col = f"{factor_group}_{letter}"
        first_level_factors[factor_col] = {
            "shared_prefix": letter,
            "source_columns": columns,
            "factor_value_logic": factor_logic,
            "signal_ls_logic": signal_logic,
            "non_null_count": int(factor_df[factor_col].notna().sum()),
        }

    source_start = None
    source_end = None
    if len(source_df.index) > 0:
        source_start = str(pd.to_datetime(source_df.index).min().date())
        source_end = str(pd.to_datetime(source_df.index).max().date())

    return {
        "factor_group": factor_group,
        "requested_factor_names": requested_factor_names,
        "source_data": {
            "input_scan": scan_record,
            "selected_columns": [str(col) for col in source_df.columns],
            "shape": list(source_df.shape),
            "index_name": source_df.index.name,
            "start_date": source_start,
            "end_date": source_end,
        },
        "grouping_rule": "Selected source factors are grouped by the first character of each factor name.",
        "first_level_factors": first_level_factors,
        "final_factor": {
            "name": factor_group,
            "source_first_level_factors": [f"{factor_group}_{letter}" for letter in groups],
            "factor_value_logic": final_logic,
            "signal_ls_logic": signal_logic,
            "non_null_count": int(factor_df[factor_group].notna().sum()),
        },
        "outputs": {
            "mounted_normalized_factors_parquet": str(output_paths["factor_parquet"]),
            "mounted_normalized_factors_xlsx": str(output_paths["factor_xlsx"]),
            "signal_ls_parquet": str(output_paths["signal_parquet"]),
            "signal_ls_xlsx": str(output_paths["signal_xlsx"]),
            "record_json": str(output_paths["record_json"]),
        },
        "warnings": warnings,
    }


# ========== 6. Interactive main ==========

def ask_factor_group() -> str:
    factor_group = input("请输入组合名称，例如 W003：").strip()

    if not factor_group:
        raise ValueError("组合名称不能为空。")

    return factor_group


def ask_factor_names() -> list[str]:
    raw = input(
        "请输入底层因子名列表，例如 L002,L003,M001，"
        "或 ['L002', 'L003', 'M001']：\n"
    )
    return parse_factor_list(raw)


def main() -> None:
    factor_group = ask_factor_group()
    factor_names = ask_factor_names()
    output_paths = build_output_paths(factor_group)
    confirm_overwrite_if_needed(output_paths)

    print("\n========== 参数确认 ==========")
    print("factor_group:", factor_group)
    print("factor_names:", factor_names)
    print("input_dir:", INPUT_DIR)
    print("output_dir:", OUTPUT_DIR)

    input_files = list_input_files(INPUT_DIR)

    print("\n========== 扫描 input_COMB ==========")
    print(f"found {len(input_files)} input files:")
    for path in input_files:
        print("-", path.name)

    source_df, scan_record, scan_warnings = collect_selected_factors(
        factor_names=factor_names,
        input_files=input_files,
    )

    factor_df, groups, factor_warnings = build_factor_matrix(
        source_df=source_df,
        factor_group=factor_group,
    )

    warnings = scan_warnings + factor_warnings
    signal_ls_df = build_signal_ls_matrix(factor_df)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    record = build_record(
        factor_group=factor_group,
        requested_factor_names=factor_names,
        source_df=source_df,
        factor_df=factor_df,
        groups=groups,
        scan_record=scan_record,
        warnings=warnings,
        output_paths=output_paths,
    )

    factor_df.to_parquet(output_paths["factor_parquet"])
    signal_ls_df.to_parquet(output_paths["signal_parquet"])

    write_factor_frame_xlsx(factor_df, output_paths["factor_xlsx"])
    write_factor_frame_xlsx(signal_ls_df, output_paths["signal_xlsx"])

    output_paths["record_json"].write_text(
        json.dumps(record, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print("\n========== 生成完成 ==========")
    print("source_df shape:", source_df.shape)
    print("factor_df shape:", factor_df.shape)
    print("signal_ls_df shape:", signal_ls_df.shape)
    print("factor columns:", list(factor_df.columns))
    print("source groups:", {letter: columns for letter, columns in groups.items()})

    print("\n========== 输出文件 ==========")
    for key, path in output_paths.items():
        print(f"{key}: {path}")

    if warnings:
        print("\n========== warnings ==========")
        for warning in warnings:
            print("-", warning)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n已取消。", file=sys.stderr)
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001
        print(f"\nERROR: {exc}", file=sys.stderr)
        sys.exit(1)
