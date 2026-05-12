"""Build W003_signal grouped factor and signal matrices from signal_ls input."""

from __future__ import annotations

import json
from collections import OrderedDict
from pathlib import Path

import pandas as pd


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[2]

FACTOR_GROUP = "W003_signal"
SOURCE_FILE_NAME = "如何从赔率和胜率看成长价值轮动——市场风格轮动系列_signal_ls.parquet"
INPUT_PATH = PROJECT_ROOT / "F_grouping" / "input_COMB" / SOURCE_FILE_NAME
OUTPUT_DIR = PROJECT_ROOT / "F_grouping" / "output_COMB"

FACTOR_OUTPUT_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_mounted_normalized_factors.parquet"
SIGNAL_OUTPUT_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_signal_ls.parquet"
FACTOR_XLSX_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_mounted_normalized_factors.xlsx"
SIGNAL_XLSX_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_signal_ls.xlsx"
RECORD_JSON_PATH = OUTPUT_DIR / f"{FACTOR_GROUP}_factor_grouping_record.json"


def group_columns_by_first_letter(columns: pd.Index) -> OrderedDict[str, list[str]]:
    groups: OrderedDict[str, list[str]] = OrderedDict()
    for column in columns:
        col = str(column)
        if not col:
            continue
        groups.setdefault(col[0], []).append(col)
    return groups


def build_factor_matrix(source_df: pd.DataFrame) -> tuple[pd.DataFrame, OrderedDict[str, list[str]], list[str]]:
    warnings: list[str] = []
    source_numeric = source_df.apply(pd.to_numeric, errors="coerce")
    groups = group_columns_by_first_letter(source_numeric.columns)

    if not groups:
        warnings.append("No non-empty source column names were found; output matrices will be empty.")

    factor_df = pd.DataFrame(index=source_numeric.index)
    for letter, columns in groups.items():
        factor_col = f"W003_{letter}_signal"
        factor_df[factor_col] = source_numeric[columns].mean(axis=1, skipna=True)
        if source_numeric[columns].notna().sum(axis=1).eq(0).any():
            empty_count = int(source_numeric[columns].notna().sum(axis=1).eq(0).sum())
            warnings.append(f"{factor_col}: {empty_count} rows have all source signals as NaN.")

    factor_df[FACTOR_GROUP] = factor_df.mean(axis=1, skipna=True)
    return factor_df.astype("float64"), groups, warnings


def build_signal_ls_matrix(factor_df: pd.DataFrame) -> pd.DataFrame:
    return factor_df.copy()


def write_factor_frame_xlsx(df: pd.DataFrame, path: Path) -> None:
    out = df.copy()
    out.index = pd.to_datetime(out.index).normalize()
    out.index.name = out.index.name or "date"
    out.to_excel(path)

    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    for cell in ws["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    wb.save(path)


def build_record(
    source_df: pd.DataFrame,
    factor_df: pd.DataFrame,
    groups: OrderedDict[str, list[str]],
    warnings: list[str],
) -> dict[str, object]:
    first_level_logic = (
        "Group source signal_ls columns by their shared first letter. For each date, compute the "
        "equal-weight row-wise mean of non-null source signals in the group; if all source signals "
        "in the group are NaN, the first-level factor is NaN. No z-score, division by 3, thresholding, "
        "or clipping is applied."
    )
    signal_logic = "signal_ls equals the factor value exactly; no additional transformation is applied."
    final_logic = (
        "W003_signal is the equal-weight row-wise mean of all first-level W003_{letter}_signal "
        "factor values, skipping NaN first-level values. No z-score, division by 3, thresholding, "
        "or clipping is applied."
    )

    first_level_factors = {}
    for letter, columns in groups.items():
        factor_col = f"W003_{letter}_signal"
        first_level_factors[factor_col] = {
            "shared_prefix": letter,
            "source_columns": columns,
            "weights": {column: 1.0 / len(columns) for column in columns},
            "factor_value_logic": first_level_logic,
            "signal_ls_logic": signal_logic,
            "non_null_count": int(factor_df[factor_col].notna().sum()),
        }

    return {
        "factor_group": FACTOR_GROUP,
        "source_data": {
            "path": str(INPUT_PATH),
            "file_name": INPUT_PATH.name,
            "columns": [str(col) for col in source_df.columns],
            "shape": list(source_df.shape),
            "index_name": source_df.index.name,
            "start_date": str(pd.to_datetime(source_df.index).min().date()),
            "end_date": str(pd.to_datetime(source_df.index).max().date()),
        },
        "first_level_factors": first_level_factors,
        "final_factor": {
            "name": FACTOR_GROUP,
            "source_first_level_factors": [f"W003_{letter}_signal" for letter in groups],
            "weights": {f"W003_{letter}_signal": 1.0 / len(groups) for letter in groups} if groups else {},
            "factor_value_logic": final_logic,
            "signal_ls_logic": signal_logic,
            "non_null_count": int(factor_df[FACTOR_GROUP].notna().sum()) if FACTOR_GROUP in factor_df else 0,
        },
        "outputs": {
            "mounted_normalized_factors_parquet": str(FACTOR_OUTPUT_PATH),
            "mounted_normalized_factors_xlsx": str(FACTOR_XLSX_PATH),
            "signal_ls_parquet": str(SIGNAL_OUTPUT_PATH),
            "signal_ls_xlsx": str(SIGNAL_XLSX_PATH),
            "record_json": str(RECORD_JSON_PATH),
        },
        "warnings": warnings,
    }


def main() -> None:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"Input signal matrix not found: {INPUT_PATH}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_df = pd.read_parquet(INPUT_PATH)
    source_df.index = pd.to_datetime(source_df.index)
    source_df = source_df.sort_index()

    factor_df, groups, warnings = build_factor_matrix(source_df)
    signal_ls_df = build_signal_ls_matrix(factor_df)
    record = build_record(source_df, factor_df, groups, warnings)

    factor_df.to_parquet(FACTOR_OUTPUT_PATH)
    signal_ls_df.to_parquet(SIGNAL_OUTPUT_PATH)
    write_factor_frame_xlsx(factor_df, FACTOR_XLSX_PATH)
    write_factor_frame_xlsx(signal_ls_df, SIGNAL_XLSX_PATH)
    RECORD_JSON_PATH.write_text(json.dumps(record, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print("factor_df shape:", factor_df.shape)
    print("signal_ls_df shape:", signal_ls_df.shape)
    print("factor columns:", list(factor_df.columns))
    print("source groups:", {letter: columns for letter, columns in groups.items()})
    print("outputs saved to:", OUTPUT_DIR)
    if warnings:
        print("warnings:")
        for warning in warnings:
            print("-", warning)


if __name__ == "__main__":
    main()
